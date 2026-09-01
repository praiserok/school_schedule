from collections import defaultdict

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.db import models
from django.forms import inlineformset_factory

from .models import Teacher, Subject, Room, SchoolClass, TeacherSubject, Schedule, Lesson, BellSchedule, BellPeriod
from .forms import (
    TeacherForm, SubjectForm, RoomForm, SchoolClassForm,
    TeacherSubjectForm, TeacherLoadRowForm, ClassLoadRowForm, ScheduleForm,
    BellScheduleForm, BellPeriodForm, DAYS_LABELS, DAYS_FULL, MAX_PERIODS,
)

DAYS_SHORT = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Нд']


# ─── Dashboard ───────────────────────────────────────────────────────────────

def dashboard(request):
    active_schedule = Schedule.objects.filter(is_active=True).first()

    from itertools import groupby
    teacher_qs = list(
        Teacher.objects.prefetch_related(
            models.Prefetch(
                'teachersubject_set',
                queryset=TeacherSubject.objects.select_related('subject').order_by('subject__name'),
            )
        ).order_by('last_name')
    )
    for t in teacher_qs:
        t.total_hours = sum(ts.hours_per_week for ts in t.teachersubject_set.all())

    def first_subject(t):
        ts = t.teachersubject_set.all()
        return ts[0].subject if ts else None

    teacher_qs.sort(key=lambda t: (
        first_subject(t).name if first_subject(t) else '\xff',
        t.last_name,
    ))

    teacher_groups = [
        (subj, list(items))
        for subj, items in groupby(teacher_qs, key=first_subject)
    ]

    return render(request, 'scheduler/dashboard.html', {
        'counts': {
            'teachers': Teacher.objects.count(),
            'subjects': Subject.objects.count(),
            'rooms': Room.objects.count(),
            'classes': SchoolClass.objects.count(),
            'assignments': TeacherSubject.objects.count(),
        },
        'active_schedule': active_schedule,
        'lesson_count': active_schedule.lessons.count() if active_schedule else 0,
        'teacher_groups': teacher_groups,
    })


# ─── Help ─────────────────────────────────────────────────────────────────────

def help_page(request):
    return render(request, 'scheduler/help.html')


# ─── Новий навчальний рік ─────────────────────────────────────────────────────

def new_year(request):
    """
    GET  → прев'ю змін (що буде видалено / зсунуто)
    POST action=reset  → повне очищення бази (залишає Teachers/Subjects/Rooms)
    POST action=shift  → зсув класів на 1 рік + нові 1-ші класи
    POST action=wipe   → видалити абсолютно все (тестові дані)
    """
    from .models import Lesson, Schedule, TeacherSubject, SchoolClass

    classes_by_grade = {}
    for cls in SchoolClass.objects.order_by('grade', 'letter'):
        classes_by_grade.setdefault(cls.grade, []).append(cls)

    max_grade = max(classes_by_grade.keys(), default=0)

    stats = {
        'lessons':     Lesson.objects.count(),
        'schedules':   Schedule.objects.count(),
        'assignments': TeacherSubject.objects.count(),
        'classes':     SchoolClass.objects.count(),
        'teachers':    Teacher.objects.count(),
        'subjects':    Subject.objects.count(),
        'rooms':       Room.objects.count(),
    }

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'wipe':
            # Видалити абсолютно всі дані
            from .models import BellSchedule
            Lesson.objects.all().delete()
            Schedule.objects.all().delete()
            TeacherSubject.objects.all().delete()
            SchoolClass.objects.all().delete()
            Teacher.objects.all().delete()
            Subject.objects.all().delete()
            Room.objects.all().delete()
            BellSchedule.objects.all().delete()
            messages.success(request, 'Базу повністю очищено.')
            return redirect('scheduler:new_year')

        if action == 'reset':
            # Очистити розклади й навантаження, залишити Teachers/Subjects/Rooms/Classes
            Lesson.objects.all().delete()
            Schedule.objects.all().delete()
            TeacherSubject.objects.all().delete()
            messages.success(request, 'Розклади та навантаження очищено.')
            return redirect('scheduler:new_year')

        if action == 'shift':
            # Зсунути паралелі на 1 рік
            new_letters = [
                l.strip().upper()
                for l in request.POST.get('new_grade1_letters', '').split(',')
                if l.strip()
            ]
            # Спочатку видаляємо випускний клас
            SchoolClass.objects.filter(grade=max_grade).delete()
            # Зсуваємо від старшого до молодшого, щоб не порушити unique_together
            for cls in SchoolClass.objects.order_by('-grade'):
                cls.grade += 1
                cls.save()
            # Очищаємо навантаження і розклади (вони більше не відповідають класам)
            Lesson.objects.all().delete()
            Schedule.objects.all().delete()
            TeacherSubject.objects.all().delete()
            # Додаємо нові 1-ші класи
            for letter in new_letters:
                SchoolClass.objects.get_or_create(grade=1, letter=letter)
            messages.success(
                request,
                f'Класи зсунуто. Видалено {max_grade}-ту паралель. '
                f'Додано нових 1-х класів: {len(new_letters)}.'
            )
            return redirect('scheduler:new_year')

    return render(request, 'scheduler/new_year.html', {
        'stats': stats,
        'classes_by_grade': dict(sorted(classes_by_grade.items())),
        'max_grade': max_grade,
    })


# ─── Teachers ────────────────────────────────────────────────────────────────

def teacher_list(request):
    subject_filter = request.GET.get('subject', '')
    search = request.GET.get('q', '').strip()
    qs = Teacher.objects.prefetch_related(
        models.Prefetch(
            'teachersubject_set',
            queryset=TeacherSubject.objects.select_related('subject').order_by('subject__name'),
        )
    )
    if subject_filter:
        qs = qs.filter(teachersubject__subject_id=subject_filter).distinct()

    from itertools import groupby
    teacher_list_data = list(qs.order_by('last_name'))
    if search:
        q = search.lower()
        teacher_list_data = [
            t for t in teacher_list_data
            if q in t.last_name.lower() or q in t.first_name.lower()
        ]

    for t in teacher_list_data:
        t.total_hours = sum(ts.hours_per_week for ts in t.teachersubject_set.all())

    def first_subject(t):
        ts = t.teachersubject_set.all()
        return ts[0].subject if ts else None

    teacher_list_data.sort(key=lambda t: (
        first_subject(t).name if first_subject(t) else '\xff',
        t.last_name,
    ))

    groups = []
    for subj, items in groupby(teacher_list_data, key=first_subject):
        teachers = list(items)
        if subj:
            subj_hours = sum(
                ts.hours_per_week
                for t in teachers
                for ts in t.teachersubject_set.all()
                if ts.subject_id == subj.pk
            )
        else:
            subj_hours = 0
        groups.append((subj, teachers, subj_hours))

    subjects = Subject.objects.order_by('name')
    active_schedule = Schedule.objects.filter(is_active=True).first()
    return render(request, 'scheduler/teacher_list.html', {
        'groups': groups,
        'subjects': subjects,
        'subject_filter': subject_filter,
        'search': search,
        'total_count': len(teacher_list_data),
        'active_schedule': active_schedule,
    })


def teacher_form(request, pk=None):
    obj = get_object_or_404(Teacher, pk=pk) if pk else None
    form = TeacherForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Збережено.')
        return redirect('scheduler:teacher_list')
    return render(request, 'scheduler/teacher_form.html', {
        'form': form,
        'title': 'Редагувати вчителя' if obj else 'Додати вчителя',
        'back_url': 'scheduler:teacher_list',
        'days_labels': DAYS_LABELS,
        'max_periods': MAX_PERIODS,
        'periods_range': range(MAX_PERIODS),
        'days_range': range(5),
    })


def teacher_delete(request, pk):
    obj = get_object_or_404(Teacher, pk=pk)
    if request.method == 'POST':
        if Lesson.objects.filter(teacher=obj).exists():
            messages.error(request, 'Неможна видалити вчителя — він присутній у згенерованому розкладі.')
            return redirect('scheduler:teacher_list')
        obj.delete()
        messages.success(request, 'Вчителя видалено.')
        return redirect('scheduler:teacher_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:teacher_list',
    })


# ─── Subjects ────────────────────────────────────────────────────────────────

def subject_list(request):
    return render(request, 'scheduler/subject_list.html', {
        'subjects': Subject.objects.all(),
    })


def subject_form(request, pk=None):
    obj = get_object_or_404(Subject, pk=pk) if pk else None
    form = SubjectForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Збережено.')
        return redirect('scheduler:subject_list')
    return render(request, 'scheduler/form.html', {
        'form': form,
        'title': 'Редагувати предмет' if obj else 'Додати предмет',
        'back_url': 'scheduler:subject_list',
    })


def subject_delete(request, pk):
    obj = get_object_or_404(Subject, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Предмет видалено.')
        return redirect('scheduler:subject_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:subject_list',
    })


# ─── Rooms ───────────────────────────────────────────────────────────────────

def room_list(request):
    active = (Schedule.objects.filter(is_active=True).first()
              or Schedule.objects.order_by('-created_at').first())
    return render(request, 'scheduler/room_list.html', {
        'rooms': Room.objects.prefetch_related('subjects').all(),
        'active_schedule': active,
    })


def room_form(request, pk=None):
    obj = get_object_or_404(Room, pk=pk) if pk else None
    form = RoomForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Збережено.')
        return redirect('scheduler:room_list')
    return render(request, 'scheduler/form.html', {
        'form': form,
        'title': 'Редагувати кабінет' if obj else 'Додати кабінет',
        'back_url': 'scheduler:room_list',
    })


def room_delete(request, pk):
    obj = get_object_or_404(Room, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Кабінет видалено.')
        return redirect('scheduler:room_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:room_list',
    })


# ─── Classes ─────────────────────────────────────────────────────────────────

def class_list(request):
    return render(request, 'scheduler/class_list.html', {
        'classes': SchoolClass.objects.select_related('home_room').all(),
    })


def class_form(request, pk=None):
    obj = get_object_or_404(SchoolClass, pk=pk) if pk else None
    form = SchoolClassForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Збережено.')
        return redirect('scheduler:class_list')
    return render(request, 'scheduler/form.html', {
        'form': form,
        'title': 'Редагувати клас' if obj else 'Додати клас',
        'back_url': 'scheduler:class_list',
    })


def class_delete(request, pk):
    obj = get_object_or_404(SchoolClass, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Клас видалено.')
        return redirect('scheduler:class_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:class_list',
    })


from django.forms import inlineformset_factory as _ifs

ClassLoadFormSet = _ifs(
    SchoolClass, TeacherSubject,
    form=ClassLoadRowForm,
    fk_name='school_class',
    extra=1,
    can_delete=True,
)


def class_load(request, pk):
    cls = get_object_or_404(SchoolClass, pk=pk)
    if request.method == 'POST':
        formset = ClassLoadFormSet(request.POST, instance=cls)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Навантаження збережено.')
            return redirect('scheduler:class_load', pk=cls.pk)
    else:
        formset = ClassLoadFormSet(instance=cls)

    # Підрахунок годин на тиждень (по слотах класу, не по кількості записів)
    all_ts = list(TeacherSubject.objects.filter(school_class=cls).order_by('subject__name', 'group'))
    seen = set()
    slot_hours = 0
    for ts in all_ts:
        key = (ts.subject_id, ts.group if ts.group else None)
        if ts.group is None or key not in seen:
            slot_hours += ts.hours_per_week
        seen.add(key)

    return render(request, 'scheduler/class_load.html', {
        'cls': cls,
        'formset': formset,
        'slot_hours': slot_hours,
    })


# ─── TeacherSubject (Навантаження) ───────────────────────────────────────────

def ts_list(request):
    qs = TeacherSubject.objects.select_related('teacher', 'subject', 'school_class').all()
    return render(request, 'scheduler/ts_list.html', {'assignments': qs})


def ts_form(request, pk=None):
    obj = get_object_or_404(TeacherSubject, pk=pk) if pk else None
    form = TeacherSubjectForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Збережено.')
        return redirect('scheduler:ts_list')
    return render(request, 'scheduler/form.html', {
        'form': form,
        'title': 'Редагувати навантаження' if obj else 'Додати навантаження',
        'back_url': 'scheduler:ts_list',
    })


def ts_delete(request, pk):
    obj = get_object_or_404(TeacherSubject, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Запис видалено.')
        return redirect('scheduler:ts_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:ts_list',
    })


TeacherLoadFormSet = inlineformset_factory(
    Teacher, TeacherSubject,
    form=TeacherLoadRowForm,
    extra=1,
    can_delete=True,
)


def teacher_load(request, teacher_pk):
    teacher = get_object_or_404(Teacher, pk=teacher_pk)
    if request.method == 'POST':
        formset = TeacherLoadFormSet(request.POST, instance=teacher)
        if formset.is_valid():
            formset.save()
            messages.success(request, 'Навантаження збережено.')
            return redirect('scheduler:teacher_load', teacher_pk=teacher.pk)
    else:
        formset = TeacherLoadFormSet(instance=teacher)
    total_hours = teacher.teachersubject_set.aggregate(
        total=models.Sum('hours_per_week')
    )['total'] or 0
    subject_totals = list(
        teacher.teachersubject_set
        .values('subject__name')
        .annotate(total=models.Sum('hours_per_week'))
        .order_by('subject__name')
    )
    return render(request, 'scheduler/teacher_load.html', {
        'teacher': teacher,
        'formset': formset,
        'total_hours': total_hours,
        'subject_totals': subject_totals,
    })


# ─── Bell schedules ──────────────────────────────────────────────────────────

BellPeriodFormSet = inlineformset_factory(
    BellSchedule, BellPeriod,
    form=BellPeriodForm,
    extra=1,
    can_delete=True,
)


def bell_list(request):
    return render(request, 'scheduler/bell_list.html', {
        'bells': BellSchedule.objects.prefetch_related('periods').all(),
    })


def bell_form(request, pk=None):
    obj = get_object_or_404(BellSchedule, pk=pk) if pk else None
    if request.method == 'POST':
        form = BellScheduleForm(request.POST, instance=obj)
        formset = BellPeriodFormSet(request.POST, instance=obj or BellSchedule())
        if form.is_valid() and formset.is_valid():
            saved = form.save()
            formset.instance = saved
            formset.save()
            messages.success(request, 'Збережено.')
            return redirect('scheduler:bell_edit', pk=saved.pk)
    else:
        form = BellScheduleForm(instance=obj)
        formset = BellPeriodFormSet(instance=obj or BellSchedule())
    return render(request, 'scheduler/bell_form.html', {
        'form': form,
        'formset': formset,
        'obj': obj,
    })


def bell_delete(request, pk):
    obj = get_object_or_404(BellSchedule, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Видалено.')
        return redirect('scheduler:bell_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:bell_list',
    })


# ─── Schedules ───────────────────────────────────────────────────────────────

def schedule_list(request):
    return render(request, 'scheduler/schedule_list.html', {
        'schedules': Schedule.objects.all(),
    })


def schedule_form(request, pk=None):
    obj = get_object_or_404(Schedule, pk=pk) if pk else None
    form = ScheduleForm(request.POST or None, instance=obj)
    if form.is_valid():
        form.save()
        messages.success(request, 'Збережено.')
        return redirect('scheduler:schedule_list')
    return render(request, 'scheduler/form.html', {
        'form': form,
        'title': 'Редагувати розклад' if obj else 'Новий розклад',
        'back_url': 'scheduler:schedule_list',
    })


def schedule_delete(request, pk):
    obj = get_object_or_404(Schedule, pk=pk)
    if request.method == 'POST':
        obj.delete()
        messages.success(request, 'Розклад видалено.')
        return redirect('scheduler:schedule_list')
    return render(request, 'scheduler/confirm_delete.html', {
        'obj': obj, 'back_url': 'scheduler:schedule_list',
    })


@require_POST
def schedule_reset_rooms(request, pk):
    """Скинути всі кабінети розкладу (встановити room=NULL)."""
    schedule = get_object_or_404(Schedule, pk=pk)
    count = schedule.lessons.exclude(room__isnull=True).update(room=None)
    messages.success(request, f'Кабінети скинуто ({count} уроків).')
    return redirect('scheduler:schedule_view', pk=pk)


@require_POST
def schedule_assign_rooms(request, pk):
    """Розставити кабінети по вже існуючому розкладу (не змінюючи слоти)."""
    from .generator import assign_rooms
    schedule = get_object_or_404(Schedule, pk=pk)
    mode = request.POST.get('mode', 'all')
    if mode not in ('all', 'specialized', 'junior', 'senior'):
        mode = 'all'
    mode_labels = {
        'all':        'Всі кабінети',
        'specialized': 'Фізкультура / Інформатика',
        'junior':     '1-4 класи',
        'senior':     '5+ класи',
    }
    assigned, eligible = assign_rooms(schedule, mode=mode)
    label = mode_labels[mode]
    messages.success(request, f'{label}: {assigned}/{eligible} уроків отримали кабінет.')
    return redirect('scheduler:schedule_view', pk=pk)


@require_POST
def schedule_copy(request, pk):
    """Копіює розклад разом з усіма уроками."""
    src = get_object_or_404(Schedule, pk=pk)

    new_schedule = Schedule.objects.create(
        name=f'{src.name} (копія)',
        days_per_week=src.days_per_week,
        lessons_per_day=src.lessons_per_day,
        bell_schedule=src.bell_schedule,
        is_active=False,
    )

    lessons = list(src.lessons.all())
    for l in lessons:
        l.pk = None
        l.schedule = new_schedule
    Lesson.objects.bulk_create(lessons)

    messages.success(request, f'Розклад скопійовано як «{new_schedule.name}».')
    return redirect('scheduler:schedule_view', pk=new_schedule.pk)


def _validate_move(schedule, lesson, new_day, new_period, swap=None):
    """Перевіряє переміщення уроку на (new_day, new_period). swap — урок з яким обмінюємось."""
    errors = []
    exclude = {lesson.pk}
    if swap:
        exclude.add(swap.pk)
    D = schedule.days_per_week
    week = lesson.week  # переміщення лише в межах одного тижня

    # 1. Доступність вчителя (день + слот)
    mask = lesson.teacher.available_days[:D].ljust(D, '0')
    if new_day >= D or mask[new_day] != '1':
        errors.append(f'Вчитель {lesson.teacher} недоступний у цей день')
    elif not lesson.teacher.is_slot_available(new_day, new_period):
        errors.append(f'Вчитель {lesson.teacher} недоступний на цьому уроці')

    # 2. Конфлікт вчителя
    if Lesson.objects.filter(
        schedule=schedule, teacher=lesson.teacher, week=week, day=new_day, period=new_period
    ).exclude(pk__in=exclude).exists():
        errors.append(f'Вчитель {lesson.teacher} вже має урок в цей час')

    # 3. Конфлікт класу (для обміну в тому ж класі слот не змінюється — пропускаємо)
    same_class_swap = swap and swap.school_class_id == lesson.school_class_id
    if not same_class_swap:
        existing_qs = Lesson.objects.filter(
            schedule=schedule, school_class=lesson.school_class, week=week, day=new_day, period=new_period
        ).exclude(pk__in=exclude)
        if lesson.group:
            # Груповий урок: дозволяємо іншу групу в тому ж слоті; забороняємо лише
            # якщо там є цілокласний урок або та сама група вже є.
            if existing_qs.filter(group__isnull=True).exists():
                errors.append(f'Клас {lesson.school_class}: цілокласний урок вже є в цей час')
            elif existing_qs.filter(group=lesson.group).exists():
                errors.append(f'Клас {lesson.school_class}: група {lesson.group} вже має урок в цей час')
        else:
            if existing_qs.exists():
                errors.append(f'Клас {lesson.school_class} вже має урок в цей час')

    # 4. Максимум уроків вчителя на день
    if lesson.day != new_day:
        cnt = Lesson.objects.filter(
            schedule=schedule, teacher=lesson.teacher, week=week, day=new_day
        ).exclude(pk__in=exclude).count() + 1
        if cnt > lesson.teacher.max_lessons_per_day:
            errors.append(f'Вчитель {lesson.teacher}: ліміт {lesson.teacher.max_lessons_per_day} ур/день буде перевищено')

    return errors


@require_POST
def lesson_move(request, pk):
    from django.http import JsonResponse
    import json
    schedule = get_object_or_404(Schedule, pk=pk)
    try:
        data = json.loads(request.body)
        lesson_id = data['lesson_id']
        new_day = int(data['new_day'])
        new_period = int(data['new_period'])
    except (json.JSONDecodeError, KeyError, ValueError, TypeError):
        return JsonResponse({'ok': False, 'errors': ['Некоректний запит']}, status=400)
    lesson = get_object_or_404(Lesson, pk=lesson_id, schedule=schedule)

    # Урок в тій самій комірці — нічого не робимо
    if lesson.day == new_day and lesson.period == new_period:
        return JsonResponse({'ok': True})

    from django.db import transaction

    old_day, old_period = lesson.day, lesson.period

    # "Family" = всі записи того ж уроку по всіх тижнях.
    # Базовий урок (xb=1) зберігається двома рядками: week=0 і week=1 в одному слоті.
    # Рухаємо одразу всю сім'ю — інакше два тижні розповзуться в різні слоти.
    lesson_family_qs = Lesson.objects.filter(
        schedule=schedule,
        school_class=lesson.school_class,
        teacher=lesson.teacher,
        subject=lesson.subject,
        group=lesson.group,
        day=old_day,
        period=old_period,
    )

    # Знаходимо "swap": якщо передано target_lid — беремо конкретний урок (для вибору
    # групи з двогрупової комірки). Інакше — перший урок у цільовій комірці.
    target_lid = data.get('target_lid')
    if target_lid:
        try:
            swap = Lesson.objects.get(pk=int(target_lid), schedule=schedule)
            if (swap.day != new_day or swap.period != new_period
                    or swap.school_class_id != lesson.school_class_id):
                return JsonResponse({'ok': False, 'errors': ['Некоректний цільовий урок']}, status=400)
        except (Lesson.DoesNotExist, ValueError, TypeError):
            return JsonResponse({'ok': False, 'errors': ['Урок не знайдено']}, status=400)
    else:
        # Якщо target_lid не вказано — це чисте переміщення (move), а не обмін (swap).
        # Урок просто переноситься в новий слот; будь-які існуючі уроки там залишаються.
        swap = None

    errors = _validate_move(schedule, lesson, new_day, new_period, swap)
    if swap:
        errors += _validate_move(schedule, swap, old_day, old_period, lesson)

    # Для звичайного переміщення (не swap) групового уроку — знаходимо сиблінг-групу
    # на тому ж слоті й рухаємо разом (group 1 + group 2 завжди переміщуються одним кроком).
    sibling_pks = []
    if lesson.group is not None and swap is None:
        sibling_qs = (Lesson.objects
                      .filter(schedule=schedule,
                              school_class=lesson.school_class,
                              subject=lesson.subject,
                              day=old_day, period=old_period)
                      .exclude(group=lesson.group)
                      .exclude(group=None))
        siblings = list(sibling_qs)
        sibling_pks = [s.pk for s in siblings]
        sibling_lesson = siblings[0] if siblings else None
        if sibling_lesson:
            sib_errors = _validate_move(schedule, sibling_lesson, new_day, new_period)
            errors += [f'Гр.{sibling_lesson.group}: {e}' for e in sib_errors]

    if errors:
        return JsonResponse({'ok': False, 'errors': errors})

    swap_family_qs = None
    if swap:
        # Сім'я swap — всі тижні тієї ж (teacher, subject, group) в цільовому слоті
        swap_family_qs = Lesson.objects.filter(
            schedule=schedule,
            school_class=lesson.school_class,
            teacher=swap.teacher,
            subject=swap.subject,
            group=swap.group,
            day=new_day,
            period=new_period,
        )

    # Зберігаємо pk ДО будь-яких оновлень — querysets ледачі, і після step 1
    # (swap → temp) фільтр по old day/period вже нічого не знайде в step 3.
    lesson_pks = list(lesson_family_qs.values_list('pk', flat=True))
    swap_pks = list(swap_family_qs.values_list('pk', flat=True)) if swap_family_qs is not None else []

    D = schedule.days_per_week
    P = schedule.lessons_per_day
    with transaction.atomic():
        if swap_pks:
            Lesson.objects.filter(pk__in=swap_pks).update(day=D, period=P)
        Lesson.objects.filter(pk__in=lesson_pks).update(day=new_day, period=new_period)
        if sibling_pks:
            Lesson.objects.filter(pk__in=sibling_pks).update(day=new_day, period=new_period)
        if swap_pks:
            Lesson.objects.filter(pk__in=swap_pks).update(day=old_day, period=old_period)

    return JsonResponse({'ok': True})


def _lesson_family(schedule, lesson):
    """Повертає (family_pks, is_regular) для уроку.

    is_regular=True означає, що урок присутній в обох тижнях (А і Б),
    тому кабінет треба перевіряти/блокувати в обох тижнях.
    """
    qs = Lesson.objects.filter(
        schedule=schedule,
        school_class=lesson.school_class,
        teacher=lesson.teacher,
        subject=lesson.subject,
        group=lesson.group,
        day=lesson.day,
        period=lesson.period,
    )
    rows = list(qs.values('pk', 'week'))
    pks = [r['pk'] for r in rows]
    weeks = {r['week'] for r in rows}
    # Regular урок має обидва тижні → займає кабінет в обох.
    # Alt урок — лише один тиждень → перевіряємо тільки свій.
    is_regular = 0 in weeks and 1 in weeks
    return pks, is_regular


def lesson_set_room(request, pk):
    from django.http import JsonResponse
    import json
    schedule = get_object_or_404(Schedule, pk=pk)

    if request.method == 'GET':
        lesson_id = request.GET.get('lesson_id')
        lesson = get_object_or_404(Lesson, pk=lesson_id, schedule=schedule)

        family_pks, is_regular = _lesson_family(schedule, lesson)

        occupancy: dict = defaultdict(int)
        slot_filter = {'schedule': schedule, 'day': lesson.day, 'period': lesson.period}
        if not is_regular:
            slot_filter['week'] = lesson.week
        for l in (Lesson.objects
                  .filter(**slot_filter)
                  .exclude(pk__in=family_pks)
                  .exclude(room__isnull=True)):
            occupancy[l.room_id] += 1

        rooms = Room.objects.order_by('name')
        available = []
        for r in rooms:
            if occupancy.get(r.pk, 0) < r.max_simultaneous:
                available.append({
                    'id': r.pk,
                    'name': r.name,
                    'current': r.pk == lesson.room_id,
                })

        return JsonResponse({
            'current_room_id': lesson.room_id,
            'current_room_name': lesson.room.name if lesson.room else None,
            'rooms': available,
        })

    try:
        data = json.loads(request.body)
        lesson_id = data['lesson_id']
        room_id = data.get('room_id')
    except (json.JSONDecodeError, KeyError, ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Некоректний запит'}, status=400)

    lesson = get_object_or_404(Lesson, pk=lesson_id, schedule=schedule)
    family_pks, is_regular = _lesson_family(schedule, lesson)

    if room_id is not None:
        room = get_object_or_404(Room, pk=room_id)
        used_filter = {'schedule': schedule, 'day': lesson.day, 'period': lesson.period, 'room': room}
        if not is_regular:
            used_filter['week'] = lesson.week
        used = (Lesson.objects
                .filter(**used_filter)
                .exclude(pk__in=family_pks)
                .count())
        if used >= room.max_simultaneous:
            return JsonResponse({'ok': False, 'error': f'Кабінет {room.name} вже зайнятий у цей урок'})
        lesson.room = room
    else:
        lesson.room = None

    lesson.save(update_fields=['room'])

    # Синхронізуємо кабінет на всю сім'ю (обидва тижні мають однаковий кабінет)
    Lesson.objects.filter(pk__in=family_pks).exclude(pk=lesson.pk).update(room=lesson.room)

    return JsonResponse({'ok': True,
                         'room_name': lesson.room.name if lesson.room else None,
                         'room_id': lesson.room_id})


@require_POST
def lesson_toggle_week(request, pk):
    """POST: змінити тиждень уроку (А↔Б). Лише для alt-уроків (week=0 або week=1)."""
    from django.http import JsonResponse
    import json
    schedule = get_object_or_404(Schedule, pk=pk)
    try:
        data = json.loads(request.body)
        lesson_id = data['lesson_id']
    except (json.JSONDecodeError, KeyError):
        return JsonResponse({'ok': False, 'error': 'Некоректний запит'}, status=400)

    lesson = get_object_or_404(Lesson, pk=lesson_id, schedule=schedule)
    new_week = 1 - lesson.week  # 0→1 або 1→0

    # Уся сім'я (той самий слот + вчитель + предмет + клас + група + тиждень)
    family_pks = list(Lesson.objects.filter(
        schedule=schedule,
        school_class=lesson.school_class,
        teacher=lesson.teacher,
        subject=lesson.subject,
        group=lesson.group,
        day=lesson.day, period=lesson.period,
        week=lesson.week,
    ).values_list('pk', flat=True))

    # Конфлікт класу в новому тижні
    if Lesson.objects.filter(
        schedule=schedule, school_class=lesson.school_class,
        day=lesson.day, period=lesson.period, week=new_week,
    ).exclude(pk__in=family_pks).exists():
        return JsonResponse({'ok': False, 'error': 'Клас вже має урок в цей час у цьому тижні'})

    # Конфлікт вчителя в новому тижні
    if Lesson.objects.filter(
        schedule=schedule, teacher=lesson.teacher,
        day=lesson.day, period=lesson.period, week=new_week,
    ).exclude(pk__in=family_pks).exists():
        return JsonResponse({'ok': False, 'error': f'Вчитель {lesson.teacher} вже зайнятий в цей час у цьому тижні'})

    Lesson.objects.filter(pk__in=family_pks).update(week=new_week)
    return JsonResponse({'ok': True})


def lesson_sibling(request, pk):
    """GET: знайти сиблінг-групу уроку, що знаходиться на ІНШОМУ слоті (ще не спарована)."""
    from django.http import JsonResponse
    schedule = get_object_or_404(Schedule, pk=pk)
    lesson_id = request.GET.get('lesson_id')
    lesson = get_object_or_404(Lesson, pk=lesson_id, schedule=schedule)

    if lesson.group is None:
        return JsonResponse({'found': False})

    sibling = (Lesson.objects
               .filter(schedule=schedule,
                       school_class=lesson.school_class,
                       subject=lesson.subject)
               .exclude(group=lesson.group)
               .exclude(group=None)
               .exclude(day=lesson.day, period=lesson.period)
               .first())

    if not sibling:
        return JsonResponse({'found': False})

    return JsonResponse({
        'found': True,
        'sibling_id': sibling.pk,
        'sibling_group': sibling.group,
        'sibling_day': sibling.day,
        'sibling_period': sibling.period,
        'sibling_day_name': DAYS_FULL[sibling.day] if sibling.day < len(DAYS_FULL) else f'День {sibling.day + 1}',
    })


@require_POST
def schedule_generate(request, pk):
    from .generator import generate
    optimize_teachers = request.POST.get('optimize_teachers') == '1'
    ok, msg = generate(pk, optimize_teachers=optimize_teachers)
    if ok:
        messages.success(request, msg)
    else:
        messages.error(request, msg)
    return redirect('scheduler:schedule_view', pk=pk)


def slot_lessons(request, pk):
    """GET: повертає всі уроки в слоті (day, period) — без кабінету і з кабінетом."""
    from django.http import JsonResponse
    schedule = get_object_or_404(Schedule, pk=pk)
    try:
        day = int(request.GET['day'])
        period = int(request.GET['period'])
    except (KeyError, ValueError):
        return JsonResponse({'error': 'bad params'}, status=400)

    # Предмети поточного кабінету (якщо передано room_id)
    room_subject_ids: set = set()
    try:
        room_id = int(request.GET['room_id'])
        room_obj = Room.objects.prefetch_related('subjects').get(pk=room_id)
        room_subject_ids = {s.pk for s in room_obj.subjects.all()}
    except (KeyError, ValueError, Room.DoesNotExist):
        pass

    lessons = (Lesson.objects
               .filter(schedule=schedule, day=day, period=period)
               .select_related('school_class', 'subject', 'teacher', 'room')
               .order_by('school_class__grade', 'school_class__letter', 'week'))

    # Групуємо по сім'ї (teacher, class, subject, group) → дедуплікуємо А/Б
    families: dict = defaultdict(list)
    for l in lessons:
        families[(l.teacher_id, l.school_class_id, l.subject_id, l.group)].append(l)

    no_room = []
    with_room = []
    for ls in families.values():
        has_room = all(x.room is not None for x in ls)
        l = ls[0] if has_room else next(x for x in ls if x.room is None)
        regular = len(ls) >= 2
        week_label = '' if regular else ('А' if l.week == 0 else 'Б')
        # match=True якщо предмет уроку входить до профілю цього кабінету
        match = bool(room_subject_ids) and l.subject_id in room_subject_ids
        entry = {
            'id': l.pk,
            'class': str(l.school_class),
            'subject': l.subject.short_name or l.subject.name,
            'color': l.subject.color,
            'teacher': str(l.teacher),
            'group': l.group,
            'week_label': week_label,
            'match': match,
        }
        if has_room:
            entry['room'] = l.room.name
            with_room.append(entry)
        else:
            no_room.append(entry)

    # Профільні предмети — вперед, решта за класом
    sort_key = lambda x: (0 if x['match'] else 1, x['class'])
    no_room.sort(key=sort_key)
    with_room.sort(key=sort_key)
    return JsonResponse({'no_room': no_room, 'with_room': with_room})


def schedule_unassigned(request, pk):
    """GET: повертає всі уроки без кабінету + вільні кімнати для кожного слоту."""
    from django.http import JsonResponse
    from django.db.models import Count
    schedule = get_object_or_404(Schedule, pk=pk)

    raw = (Lesson.objects
           .filter(schedule=schedule, room__isnull=True)
           .select_related('school_class', 'subject', 'teacher')
           .order_by('day', 'period', 'school_class__grade', 'school_class__letter'))

    families: dict = defaultdict(list)
    for l in raw:
        families[(l.day, l.period, l.school_class_id, l.subject_id, l.group)].append(l)

    if not families:
        return JsonResponse({'items': []})

    # Зайнятість кабінетів агрегується в БД (GROUP BY замість Python-підрахунку)
    occupancy: dict = defaultdict(lambda: defaultdict(int))
    for row in (Lesson.objects
                .filter(schedule=schedule, room__isnull=False)
                .values('day', 'period', 'week', 'room_id')
                .annotate(cnt=Count('pk'))):
        occupancy[(row['day'], row['period'], row['week'])][row['room_id']] = row['cnt']

    rooms = list(Room.objects.only('pk', 'name', 'max_simultaneous').order_by('name'))

    items = []
    for (day, period, cls_id, subj_id, group), ls in families.items():
        rep = ls[0]
        missing_weeks = {l.week for l in ls}

        free_rooms = []
        for r in rooms:
            if all(occupancy[(day, period, w)].get(r.pk, 0) < r.max_simultaneous
                   for w in missing_weeks):
                free_rooms.append({'id': r.pk, 'name': r.name})

        items.append({
            'lid':      rep.pk,
            'day':      day,
            'day_name': DAYS_SHORT[day] if day < len(DAYS_SHORT) else str(day + 1),
            'period':   period + 1,
            'class':    str(rep.school_class),
            'subject':  rep.subject.short_name or rep.subject.name,
            'color':    rep.subject.color,
            'teacher':  str(rep.teacher),
            'group':    rep.group or '',
            'free_rooms': free_rooms,
        })

    return JsonResponse({'items': items})


def schedule_unassigned_count(request, pk):
    """GET: повертає лише кількість унікальних сімей без кабінету (для бейджа)."""
    from django.http import JsonResponse
    schedule = get_object_or_404(Schedule, pk=pk)
    count = (Lesson.objects
             .filter(schedule=schedule, room__isnull=True)
             .values('day', 'period', 'school_class_id', 'subject_id', 'group')
             .distinct()
             .count())
    return JsonResponse({'count': count})


def room_schedule(request, schedule_pk, room_pk):
    schedule = get_object_or_404(Schedule, pk=schedule_pk)
    room = get_object_or_404(Room, pk=room_pk)
    lessons = (schedule.lessons.filter(room=room)
               .select_related('school_class', 'subject', 'teacher')
               .order_by('day', 'period', 'school_class__grade', 'school_class__letter', 'group', 'week'))
    days = DAYS_FULL[:schedule.days_per_week]
    D = schedule.days_per_week
    P = schedule.lessons_per_day
    periods = range(P)

    # Collect raw lessons per slot
    raw: dict = defaultdict(list)
    for lesson in lessons:
        if lesson.day < D and lesson.period < P:
            raw[lesson.day, lesson.period].append(lesson)

    # grid[d][p] → list of entries, one per class in the slot.
    # Each entry: {'kind': 'regular'|'alt'|'week_a'|'week_b',
    #              'lesson': l,          ← for regular/week_a/week_b
    #              'week_a': l|None,     ← for alt
    #              'week_b': l|None}     ← for alt
    grid = {d: {p: [] for p in range(P)} for d in range(D)}
    for d in range(D):
        for p in range(P):
            cell = raw.get((d, p), [])
            if not cell:
                continue
            by_group: dict = defaultdict(list)
            for l in cell:
                by_group[(l.school_class_id, l.group)].append(l)
            entries = []
            for ls in by_group.values():
                week_a = next((l for l in ls if l.week == 0), None)
                week_b = next((l for l in ls if l.week == 1), None)
                if week_a and week_b and week_a.subject_id == week_b.subject_id:
                    entries.append({'kind': 'regular', 'lesson': week_a})
                elif week_a and week_b:
                    entries.append({'kind': 'alt', 'week_a': week_a, 'week_b': week_b})
                elif week_a:
                    entries.append({'kind': 'week_a', 'lesson': week_a})
                else:
                    entries.append({'kind': 'week_b', 'lesson': week_b})
            grid[d][p] = entries

    bell_times = {}
    if schedule.bell_schedule_id:
        bell_times = {
            bp.number - 1: bp
            for bp in BellPeriod.objects.filter(bell_schedule_id=schedule.bell_schedule_id)
        }

    all_rooms = Room.objects.prefetch_related('subjects').all()
    all_schedules = Schedule.objects.order_by('-created_at')
    all_teachers = (Teacher.objects
                    .filter(lesson__schedule=schedule)
                    .distinct()
                    .order_by('last_name', 'first_name'))
    return render(request, 'scheduler/room_schedule.html', {
        'schedule': schedule,
        'room': room,
        'all_rooms': all_rooms,
        'all_schedules': all_schedules,
        'all_teachers': all_teachers,
        'days': days,
        'periods': periods,
        'grid': grid,
        'bell_times': bell_times,
    })


@require_POST
def assign_teacher_to_room(request, pk):
    """POST JSON {teacher_id, room_id}: призначити всі уроки вчителя в цей кабінет."""
    from django.http import JsonResponse
    import json
    schedule = get_object_or_404(Schedule, pk=pk)
    try:
        data = json.loads(request.body)
        teacher_id = int(data['teacher_id'])
        room_id = int(data['room_id'])
    except (json.JSONDecodeError, KeyError, ValueError, TypeError):
        return JsonResponse({'ok': False, 'error': 'Некоректний запит'}, status=400)
    teacher = get_object_or_404(Teacher, pk=teacher_id)
    room = get_object_or_404(Room, pk=room_id)
    count = schedule.lessons.filter(teacher=teacher).update(room=room)
    return JsonResponse({'ok': True, 'count': count, 'teacher': str(teacher), 'room': room.name})


def teacher_schedule(request, schedule_pk, teacher_pk):
    schedule = get_object_or_404(Schedule, pk=schedule_pk)
    teacher = get_object_or_404(Teacher, pk=teacher_pk)
    lessons = schedule.lessons.filter(teacher=teacher).select_related('school_class', 'subject', 'room')
    days = DAYS_FULL[:schedule.days_per_week]
    D = schedule.days_per_week
    P = schedule.lessons_per_day
    periods = range(P)

    # grid[d][p]: None | {'kind':'regular','lesson':l} | {'kind':'alt','week_a':l,'week_b':l}
    raw = {d: {p: [] for p in range(P)} for d in range(D)}
    for lesson in lessons:
        if lesson.day < D and lesson.period < P:
            raw[lesson.day][lesson.period].append(lesson)

    grid = {}
    for d in range(D):
        grid[d] = {}
        for p in range(P):
            cell = raw[d][p]
            if not cell:
                grid[d][p] = None
                continue
            week_a = [l for l in cell if l.week == 0]
            week_b = [l for l in cell if l.week == 1]
            subj_a = {l.subject_id for l in week_a}
            subj_b = {l.subject_id for l in week_b}
            if week_a and week_b and subj_a == subj_b \
                    and week_a[0].school_class_id == week_b[0].school_class_id:
                grid[d][p] = {'kind': 'regular', 'lesson': week_a[0]}
            elif week_a and week_b:
                grid[d][p] = {'kind': 'alt', 'week_a': week_a[0], 'week_b': week_b[0]}
            elif week_a:
                grid[d][p] = {'kind': 'week_a', 'lesson': week_a[0]}
            else:
                grid[d][p] = {'kind': 'week_b', 'lesson': week_b[0]}

    bell_times = {}
    if schedule.bell_schedule_id:
        bell_times = {
            bp.number - 1: bp
            for bp in BellPeriod.objects.filter(bell_schedule_id=schedule.bell_schedule_id)
        }

    all_teachers = Teacher.objects.all()
    all_schedules = Schedule.objects.order_by('-created_at')
    return render(request, 'scheduler/teacher_schedule.html', {
        'schedule': schedule,
        'teacher': teacher,
        'all_teachers': all_teachers,
        'all_schedules': all_schedules,
        'days': days,
        'periods': periods,
        'grid': grid,
        'bell_times': bell_times,
    })


def _build_display_grid(lessons, classes, D, P):
    """
    Повертає grid[class_pk][day][period] — dict з ключами:
      None                          → порожня комірка
      {'kind':'regular', 'primary': lesson, 'extra': lesson|None}
                                    → звичайний урок (або два записи однієї групи)
      {'kind':'alt', 'week_a': lesson|None, 'week_b': lesson|None}
                                    → уроки що чергуються по тижнях
    """
    raw = {sc.pk: {d: {p: [] for p in range(P)} for d in range(D)} for sc in classes}
    for lesson in lessons:
        if lesson.day < D and lesson.period < P:
            raw[lesson.school_class_id][lesson.day][lesson.period].append(lesson)

    grid = {}
    for sc in classes:
        grid[sc.pk] = {}
        for d in range(D):
            grid[sc.pk][d] = {}
            for p in range(P):
                cell = raw[sc.pk][d][p]
                if not cell:
                    grid[sc.pk][d][p] = None
                    continue
                week_a = [l for l in cell if l.week == 0]
                week_b = [l for l in cell if l.week == 1]
                subj_a = {l.subject_id for l in week_a}
                subj_b = {l.subject_id for l in week_b}
                if week_a and week_b and subj_a == subj_b:
                    primary = week_a[0]
                    extra = week_a[1] if len(week_a) > 1 else None
                    grid[sc.pk][d][p] = {'kind': 'regular', 'primary': primary, 'extra': extra}
                else:
                    # Або лише тиждень А (черговий 0.5г), або різні предмети → alt
                    grid[sc.pk][d][p] = {
                        'kind': 'alt',
                        'week_a': week_a[0] if week_a else None,
                        'week_b': week_b[0] if week_b else None,
                    }
    return grid


def schedule_view(request, pk):
    schedule = get_object_or_404(Schedule, pk=pk)
    lessons = schedule.lessons.select_related('school_class', 'subject', 'teacher', 'room').order_by('week', 'group')
    classes = SchoolClass.objects.all()
    D = schedule.days_per_week
    P = schedule.lessons_per_day
    days = DAYS_FULL[:D]
    periods = range(P)

    grid = _build_display_grid(lessons, classes, D, P)

    bell_times = {}
    if schedule.bell_schedule_id:
        bell_times = {
            bp.number - 1: bp
            for bp in BellPeriod.objects.filter(bell_schedule_id=schedule.bell_schedule_id)
        }

    return render(request, 'scheduler/schedule_view.html', {
        'schedule': schedule,
        'classes': classes,
        'days': days,
        'periods': periods,
        'grid': grid,
        'all_teachers': Teacher.objects.all(),
        'bell_times': bell_times,
    })


# ─── Оновлення вчителів у розкладі ──────────────────────────────────────────

def sync_schedule_teachers(request, pk):
    """
    GET  — показує які уроки змінять вчителя (порівняння з поточними TeacherSubject).
    POST — застосовує зміни одним SQL CASE-запитом (атомарно, без конфліктів).
    """
    from django.db import transaction

    schedule = get_object_or_404(Schedule, pk=pk)

    # Словник: (school_class_id, subject_id, group) → teacher
    ts_map = {
        (ts.school_class_id, ts.subject_id, ts.group): ts.teacher
        for ts in TeacherSubject.objects.select_related('teacher')
    }

    # Знаходимо уроки з розбіжністю
    lessons = list(
        schedule.lessons.select_related('school_class', 'subject', 'teacher')
    )
    changes = []  # [{lesson, old_teacher, new_teacher}]
    for lesson in lessons:
        key = (lesson.school_class_id, lesson.subject_id, lesson.group)
        new_teacher = ts_map.get(key)
        if new_teacher and new_teacher.pk != lesson.teacher_id:
            changes.append({
                'lesson':      lesson,
                'old_teacher': lesson.teacher,
                'new_teacher': new_teacher,
            })

    if request.method == 'POST' and changes:
        with transaction.atomic():
            # Конфліктні зміни: новий вчитель вже зайнятий в тому ж слоті в БД
            conflict_ids = set()
            for c in changes:
                occupied = Lesson.objects.filter(
                    schedule_id=schedule.pk,
                    teacher_id=c['new_teacher'].pk,
                    day=c['lesson'].day,
                    period=c['lesson'].period,
                    week=c['lesson'].week,
                ).exclude(pk=c['lesson'].pk).exists()
                if occupied:
                    conflict_ids.add(c['lesson'].pk)

            # Додатково: якщо два «безпечних» оновлення ведуть одного вчителя
            # в один слот — між собою вони теж конфліктують (один витіснить другий).
            from collections import defaultdict
            slot_map = defaultdict(list)
            for c in changes:
                if c['lesson'].pk not in conflict_ids:
                    slot = (c['new_teacher'].pk, c['lesson'].day, c['lesson'].period, c['lesson'].week)
                    slot_map[slot].append(c['lesson'].pk)
            for pks in slot_map.values():
                if len(pks) > 1:
                    conflict_ids.update(pks)

            safe        = [c for c in changes if c['lesson'].pk not in conflict_ids]
            conflicting = [c for c in changes if c['lesson'].pk in conflict_ids]

            # Безпечні — прості update (кожен новий вчитель унікальний у слоті)
            for c in safe:
                Lesson.objects.filter(pk=c['lesson'].pk).update(teacher_id=c['new_teacher'].pk)

            # Конфліктні (swap тощо) — видалити всі, потім перестворити
            if conflicting:
                new_lessons = []
                for c in conflicting:
                    l = c['lesson']
                    new_lessons.append(Lesson(
                        schedule_id=l.schedule_id,
                        school_class_id=l.school_class_id,
                        subject_id=l.subject_id,
                        teacher_id=c['new_teacher'].pk,
                        room_id=l.room_id,
                        day=l.day,
                        period=l.period,
                        week=l.week,
                        group=l.group,
                        is_double=l.is_double,
                    ))
                Lesson.objects.filter(pk__in=[c['lesson'].pk for c in conflicting]).delete()
                Lesson.objects.bulk_create(new_lessons)

        messages.success(request, f'Оновлено вчителів у {len(changes)} уроках.')
        return redirect('scheduler:schedule_view', pk=pk)

    return render(request, 'scheduler/sync_teachers.html', {
        'schedule': schedule,
        'changes':  changes,
    })


# ─── Спільні хелпери для публічних вʼюв ─────────────────────────────────────

def _get_active_schedule():
    """Повертає активний розклад або кидає Http404. Безпечно при >1 активному."""
    from django.http import Http404
    schedule = Schedule.objects.filter(is_active=True).order_by('-pk').first()
    if not schedule:
        raise Http404
    return schedule


def _get_bell_times(schedule):
    """Повертає {period_index: BellPeriod} для розкладу дзвінків."""
    if not schedule.bell_schedule_id:
        return {}
    return {bp.number - 1: bp
            for bp in BellPeriod.objects.filter(bell_schedule_id=schedule.bell_schedule_id)}


def _build_teacher_grid(lessons, D, P):
    """
    grid[d][p] → None | {'kind': 'regular'|'alt'|'only_a'|'only_b', 'lessons': [...]}
    або {'kind': 'alt', 'week_a': [...], 'week_b': [...]}.
    Вважає 'regular' тільки якщо той самий клас і предмет в обидва тижні.
    """
    raw = {d: {p: [] for p in range(P)} for d in range(D)}
    for l in lessons:
        if l.day < D and l.period < P:
            raw[l.day][l.period].append(l)
    grid = {}
    for d in range(D):
        grid[d] = {}
        for p in range(P):
            cell = raw[d][p]
            if not cell:
                grid[d][p] = None
                continue
            week_a = [l for l in cell if l.week == 0]
            week_b = [l for l in cell if l.week == 1]
            subj_a = {l.subject_id for l in week_a}
            subj_b = {l.subject_id for l in week_b}
            same_class = (week_a and week_b
                          and week_a[0].school_class_id == week_b[0].school_class_id)
            if week_a and week_b and subj_a == subj_b and same_class:
                grid[d][p] = {'kind': 'regular', 'lessons': week_a}
            elif week_a and week_b:
                grid[d][p] = {'kind': 'alt', 'week_a': week_a, 'week_b': week_b}
            elif week_a:
                grid[d][p] = {'kind': 'only_a', 'lessons': week_a}
            else:
                grid[d][p] = {'kind': 'only_b', 'lessons': week_b}
    return grid


def _build_room_entries_grid(lessons, D, P):
    """
    grid[d][p] → list of entry dicts (може бути декілька класів в одному кабінеті).
    entry: {'kind': 'regular'|'alt'|'week_a'|'week_b', 'lesson': l}
           або {'kind': 'alt', 'week_a': l, 'week_b': l}
    """
    raw: dict = defaultdict(list)
    for l in lessons:
        if l.day < D and l.period < P:
            raw[l.day, l.period].append(l)
    grid = {d: {p: [] for p in range(P)} for d in range(D)}
    for d in range(D):
        for p in range(P):
            cell = raw.get((d, p), [])
            if not cell:
                continue
            by_group: dict = defaultdict(list)
            for l in cell:
                by_group[(l.school_class_id, l.group)].append(l)
            entries = []
            for ls in by_group.values():
                week_a = next((l for l in ls if l.week == 0), None)
                week_b = next((l for l in ls if l.week == 1), None)
                if week_a and week_b and week_a.subject_id == week_b.subject_id:
                    entries.append({'kind': 'regular', 'lesson': week_a})
                elif week_a and week_b:
                    entries.append({'kind': 'alt', 'week_a': week_a, 'week_b': week_b})
                elif week_a:
                    entries.append({'kind': 'week_a', 'lesson': week_a})
                else:
                    entries.append({'kind': 'week_b', 'lesson': week_b})
            grid[d][p] = entries
    return grid


# ─── Публічний перегляд (без логіну) ────────────────────────────────────────

def public_home(request):
    """Головна публічна сторінка: вибір класу або вчителя."""
    schedule = Schedule.objects.filter(is_active=True).first()
    if not schedule:
        return render(request, 'scheduler/public_home.html', {'schedule': None})
    classes  = SchoolClass.objects.order_by('grade', 'letter')
    teachers = Teacher.objects.order_by('last_name', 'first_name')
    rooms    = Room.objects.order_by('name')
    return render(request, 'scheduler/public_home.html', {
        'schedule': schedule,
        'classes':  classes,
        'teachers': teachers,
        'rooms':    rooms,
    })


def public_class(request, pk):
    """Публічний розклад класу (тільки активний розклад)."""
    schedule     = _get_active_schedule()
    school_class = get_object_or_404(SchoolClass, pk=pk)
    D, P = schedule.days_per_week, schedule.lessons_per_day

    lessons = (schedule.lessons
               .filter(school_class=school_class)
               .select_related('subject', 'teacher', 'room')
               .order_by('week', 'group'))

    return render(request, 'scheduler/public_class.html', {
        'schedule':     schedule,
        'school_class': school_class,
        'grid':         _build_display_grid(lessons, [school_class], D, P),
        'days':         DAYS_FULL[:D],
        'periods':      range(P),
        'bell_times':   _get_bell_times(schedule),
        'classes':      SchoolClass.objects.order_by('grade', 'letter'),
        'teachers':     Teacher.objects.order_by('last_name', 'first_name'),
        'rooms':        Room.objects.only('pk', 'name').order_by('name'),
    })


def public_teacher(request, pk):
    """Публічний розклад вчителя (тільки активний розклад)."""
    schedule = _get_active_schedule()
    teacher  = get_object_or_404(Teacher, pk=pk)
    D, P = schedule.days_per_week, schedule.lessons_per_day

    lessons = (schedule.lessons
               .filter(teacher=teacher)
               .select_related('school_class', 'subject', 'room')
               .order_by('day', 'period', 'week'))

    return render(request, 'scheduler/public_teacher.html', {
        'schedule':   schedule,
        'teacher':    teacher,
        'grid':       _build_teacher_grid(lessons, D, P),
        'days':       DAYS_FULL[:D],
        'periods':    range(P),
        'bell_times': _get_bell_times(schedule),
        'classes':    SchoolClass.objects.order_by('grade', 'letter'),
        'teachers':   Teacher.objects.order_by('last_name', 'first_name'),
        'rooms':      Room.objects.only('pk', 'name').order_by('name'),
    })


def public_room(request, pk):
    """Публічний розклад кабінету (тільки активний розклад)."""
    schedule = _get_active_schedule()
    room     = get_object_or_404(Room, pk=pk)
    D, P = schedule.days_per_week, schedule.lessons_per_day

    lessons = (schedule.lessons.filter(room=room)
               .select_related('school_class', 'subject', 'teacher')
               .order_by('day', 'period', 'school_class__grade', 'school_class__letter', 'group', 'week'))

    return render(request, 'scheduler/public_room.html', {
        'schedule':   schedule,
        'room':       room,
        'grid':       _build_room_entries_grid(lessons, D, P),
        'days':       DAYS_FULL[:D],
        'periods':    range(P),
        'bell_times': _get_bell_times(schedule),
        'classes':    SchoolClass.objects.order_by('grade', 'letter'),
        'teachers':   Teacher.objects.order_by('last_name', 'first_name'),
        'rooms':      Room.objects.only('pk', 'name').order_by('name'),
    })


# ─── Public XLSX exports (без логіну) ────────────────────────────────────────

def _pub_xlsx_response(title, D, P, days_full, bell_times, get_cell, filename):
    """
    Будує XLSX-файл і повертає HttpResponse.
    get_cell(d, p) → (text: str, color_hex: str|None)
      text — багаторядковий рядок (\\n як роздільник)
      color_hex — HEX колір предмету (#rrggbb) або None
    """
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_NAVY    = '1B3A6B'
    C_BLUE    = '2C5F8A'
    C_LT_BLUE = 'D6E4F0'
    C_GRAY    = 'F0F0F0'
    C_WHITE   = 'FFFFFF'

    thin    = Side(style='thin', color='BFBFBF')
    brd     = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_hdr = Border(left=thin, right=thin, top=thin,
                     bottom=Side(style='medium', color=C_NAVY))

    f_title  = Font(name='Calibri', bold=True, size=14, color=C_WHITE)
    f_day    = Font(name='Calibri', bold=True, size=11, color=C_WHITE)
    f_period = Font(name='Calibri', bold=True, size=10, color=C_NAVY)
    f_cell   = Font(name='Calibri', size=10)

    fill_navy  = PatternFill('solid', fgColor=C_NAVY)
    fill_blue  = PatternFill('solid', fgColor=C_BLUE)
    fill_lblue = PatternFill('solid', fgColor=C_LT_BLUE)
    fill_gray  = PatternFill('solid', fgColor=C_GRAY)

    al_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_l = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _subject_fill(hex_color, alpha=0.22):
        try:
            h = (hex_color or '').lstrip('#')
            if len(h) != 6:
                return PatternFill('solid', fgColor=C_WHITE)
            r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
            r2 = round(r * alpha + 255 * (1 - alpha))
            g2 = round(g * alpha + 255 * (1 - alpha))
            b2 = round(b * alpha + 255 * (1 - alpha))
            return PatternFill('solid', fgColor=f'{r2:02X}{g2:02X}{b2:02X}')
        except Exception:
            return PatternFill('solid', fgColor=C_WHITE)

    LINE_H  = 18
    ROW_MIN = 52

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.column_dimensions['A'].width = 13
    for i in range(D):
        ws.column_dimensions[get_column_letter(i + 2)].width = 30

    # Рядок 1 — заголовок
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=D + 1)
    c = ws.cell(row=1, column=1, value=title)
    c.font = f_title; c.fill = fill_navy; c.alignment = al_c
    ws.row_dimensions[1].height = 26

    # Рядок 2 — дні тижня
    hdr = ws.cell(row=2, column=1, value='Урок')
    hdr.font = f_day; hdr.fill = fill_blue; hdr.alignment = al_c; hdr.border = brd_hdr
    for i, day in enumerate(days_full):
        c = ws.cell(row=2, column=i + 2, value=day)
        c.font = f_day; c.fill = fill_blue; c.alignment = al_c; c.border = brd_hdr
    ws.row_dimensions[2].height = 18

    # Рядки 3+ — уроки
    for p in range(P):
        row = p + 3
        bp = bell_times.get(p)
        period_val = (f'{p + 1}\n{bp.start_time.strftime("%H:%M")}–{bp.end_time.strftime("%H:%M")}'
                      if bp else str(p + 1))
        pc = ws.cell(row=row, column=1, value=period_val)
        pc.font = f_period; pc.fill = fill_lblue; pc.alignment = al_c; pc.border = brd

        max_lines = 1
        for d in range(D):
            text, color = get_cell(d, p)
            c = ws.cell(row=row, column=d + 2)
            c.border = brd; c.font = f_cell
            if text:
                c.value = text
                c.alignment = al_l
                c.fill = _subject_fill(color) if color else PatternFill('solid', fgColor=C_WHITE)
                max_lines = max(max_lines, text.count('\n') + 1)
            else:
                c.fill = fill_gray
                c.alignment = al_c

        ws.row_dimensions[row].height = max(max_lines * LINE_H, ROW_MIN)

    ws.freeze_panes = 'B3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return resp


def public_class_export(request, pk):
    """XLSX розклад класу (публічний)."""
    schedule     = _get_active_schedule()
    school_class = get_object_or_404(SchoolClass, pk=pk)
    D, P = schedule.days_per_week, schedule.lessons_per_day

    lessons = (schedule.lessons.filter(school_class=school_class)
               .select_related('subject', 'teacher', 'room')
               .order_by('week', 'group'))
    grid = _build_display_grid(lessons, [school_class], D, P)[school_class.pk]
    bell_times = _get_bell_times(schedule)

    def _lesson_line(l):
        room = f' · каб. {l.room}' if l.room else ''
        return f'{l.teacher}{room}'

    def get_cell(d, p):
        cell = grid[d][p]
        if cell is None:
            return '', None
        if cell['kind'] == 'regular':
            l = cell['primary']
            subj = l.subject.short_name or l.subject.name
            lines = [subj]
            if cell['extra']:
                ex = cell['extra']
                g1 = l.group or '1'
                g2 = ex.group or '2'
                lines.append(f'гр.{g1}: {_lesson_line(l)}')
                lines.append(f'гр.{g2}: {_lesson_line(ex)}')
            else:
                lines.append(_lesson_line(l))
            return '\n'.join(lines), l.subject.color
        # alt
        parts = []
        if cell['week_a']:
            la = cell['week_a']
            sa = la.subject.short_name or la.subject.name
            grp = f' гр.{la.group}' if la.group else ''
            parts.append(f'А: {sa}{grp}\n{_lesson_line(la)}')
        if cell['week_b']:
            lb = cell['week_b']
            sb = lb.subject.short_name or lb.subject.name
            grp = f' гр.{lb.group}' if lb.group else ''
            parts.append(f'Б: {sb}{grp}\n{_lesson_line(lb)}')
        color = (cell['week_a'] or cell['week_b']).subject.color
        return '\n'.join(parts), color

    safe = str(school_class).replace(' ', '_')
    return _pub_xlsx_response(
        title=f'Розклад {school_class}  ·  {schedule.name}',
        D=D, P=P, days_full=DAYS_FULL[:D], bell_times=bell_times,
        get_cell=get_cell, filename=f'schedule_class_{safe}',
    )


def public_teacher_export(request, pk):
    """XLSX розклад вчителя (публічний)."""
    schedule = _get_active_schedule()
    teacher  = get_object_or_404(Teacher, pk=pk)
    D, P = schedule.days_per_week, schedule.lessons_per_day

    lessons = (schedule.lessons.filter(teacher=teacher)
               .select_related('school_class', 'subject', 'room')
               .order_by('day', 'period', 'week'))

    grid = _build_teacher_grid(lessons, D, P)
    bell_times = _get_bell_times(schedule)

    def _cls_str(l):
        return str(l.school_class) + (f' гр.{l.group}' if l.group else '')

    def get_cell(d, p):
        cell = grid[d][p]
        if cell is None:
            return '', None
        if cell['kind'] in ('regular', 'only_a', 'only_b'):
            ls = cell['lessons']
            l0 = ls[0]
            subj = l0.subject.short_name or l0.subject.name
            classes = ', '.join(_cls_str(l) for l in ls)
            lines = [subj, classes]
            if l0.room:
                lines.append(f'каб. {l0.room}')
            return '\n'.join(lines), l0.subject.color
        # alt
        parts = []
        if cell['week_a']:
            la = cell['week_a'][0]
            sa = la.subject.short_name or la.subject.name
            classes_a = ', '.join(_cls_str(l) for l in cell['week_a'])
            parts.append(f'А: {sa}\n{classes_a}' + (f'\nкаб. {la.room}' if la.room else ''))
        if cell['week_b']:
            lb = cell['week_b'][0]
            sb = lb.subject.short_name or lb.subject.name
            classes_b = ', '.join(_cls_str(l) for l in cell['week_b'])
            parts.append(f'Б: {sb}\n{classes_b}' + (f'\nкаб. {lb.room}' if lb.room else ''))
        color = (cell['week_a'][0] if cell['week_a'] else cell['week_b'][0]).subject.color
        return '\n'.join(parts), color

    safe = str(teacher).replace(' ', '_')[:30]
    return _pub_xlsx_response(
        title=f'{teacher}  ·  {schedule.name}',
        D=D, P=P, days_full=DAYS_FULL[:D], bell_times=bell_times,
        get_cell=get_cell, filename=f'schedule_teacher_{safe}',
    )


def public_room_export(request, pk):
    """XLSX навантаженість кабінету (публічний)."""
    schedule = _get_active_schedule()
    room     = get_object_or_404(Room, pk=pk)
    D, P = schedule.days_per_week, schedule.lessons_per_day

    lessons = (schedule.lessons.filter(room=room)
               .select_related('school_class', 'subject', 'teacher')
               .order_by('day', 'period', 'school_class__grade', 'school_class__letter', 'group', 'week'))

    grid = _build_room_entries_grid(lessons, D, P)
    bell_times = _get_bell_times(schedule)

    def get_cell(d, p):
        entries = grid[d][p]
        if not entries:
            return '', None
        lines = []
        first_color = None
        for i, entry in enumerate(entries):
            if i > 0:
                lines.append('─────')
            if entry['kind'] in ('regular', 'week_a', 'week_b'):
                l = entry['lesson']
                subj = l.subject.short_name or l.subject.name
                grp = f' гр.{l.group}' if l.group else ''
                prefix = {'week_a': 'А: ', 'week_b': 'Б: '}.get(entry['kind'], '')
                lines += [f'{prefix}{subj}', f'{l.school_class}{grp}', str(l.teacher)]
                if first_color is None:
                    first_color = l.subject.color
            else:  # alt
                la, lb = entry.get('week_a'), entry.get('week_b')
                if la:
                    sa = la.subject.short_name or la.subject.name
                    grp = f' гр.{la.group}' if la.group else ''
                    lines += [f'А: {sa}', f'{la.school_class}{grp}', str(la.teacher)]
                    if first_color is None:
                        first_color = la.subject.color
                if lb:
                    sb = lb.subject.short_name or lb.subject.name
                    grp = f' гр.{lb.group}' if lb.group else ''
                    lines += [f'Б: {sb}', f'{lb.school_class}{grp}', str(lb.teacher)]
        return '\n'.join(lines), first_color

    safe = room.name.replace(' ', '_')[:30]
    return _pub_xlsx_response(
        title=f'Кабінет {room.name}  ·  {schedule.name}',
        D=D, P=P, days_full=DAYS_FULL[:D], bell_times=bell_times,
        get_cell=get_cell, filename=f'schedule_room_{safe}',
    )


# ─── Exports ─────────────────────────────────────────────────────────────────

def export_teacher_load(request):
    """XLSX: all teachers and their subject load per class."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Навантаження вчителів'

    # ── Styles ──
    hdr_font   = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill   = PatternFill('solid', fgColor='2C5F8A')
    sub_font   = Font(bold=True, size=10)
    sub_fill   = PatternFill('solid', fgColor='D6E4F0')
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill('solid', fgColor='EAF4EA')
    center     = Alignment(horizontal='center', vertical='center')
    thin       = Side(style='thin', color='BFBFBF')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value, font=None, fill=None, align=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:   c.font      = font
        if fill:   c.fill      = fill
        if align:  c.alignment = align
        if num_fmt: c.number_format = num_fmt
        c.border = border
        return c

    # ── Header row ──
    headers = ['Вчитель', 'Предмет', 'Клас', 'Група', 'Год/тиж']
    for col, h in enumerate(headers, 1):
        cell(1, col, h, font=hdr_font, fill=hdr_fill, align=center)
    ws.row_dimensions[1].height = 20

    # ── Data ──
    assignments = (
        TeacherSubject.objects
        .select_related('teacher', 'subject', 'school_class')
        .order_by('teacher__last_name', 'teacher__first_name',
                  'school_class__grade', 'school_class__letter',
                  'subject__name', 'group')
    )

    row = 2
    current_teacher = None
    teacher_start   = 2
    teacher_total   = 0

    for ts in assignments:
        teacher_name = str(ts.teacher)

        # Teacher subtotal row when teacher changes
        if current_teacher is not None and teacher_name != current_teacher:
            cell(row, 1, f'Разом: {current_teacher}', font=total_font, fill=total_fill)
            cell(row, 2, '', fill=total_fill)
            cell(row, 3, '', fill=total_fill)
            cell(row, 4, '', fill=total_fill)
            cell(row, 5, teacher_total, font=total_font, fill=total_fill,
                 align=center, num_fmt='0.0')
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1
            teacher_total = 0

        current_teacher = teacher_name
        teacher_total  += ts.hours_per_week

        cell(row, 1, teacher_name)
        cell(row, 2, ts.subject.name)
        cell(row, 3, str(ts.school_class), align=center)
        cell(row, 4, ts.group if ts.group else '', align=center)
        cell(row, 5, float(ts.hours_per_week), align=center, num_fmt='0.0')
        row += 1

    # Last teacher subtotal
    if current_teacher:
        cell(row, 1, f'Разом: {current_teacher}', font=total_font, fill=total_fill)
        cell(row, 2, '', fill=total_fill)
        cell(row, 3, '', fill=total_fill)
        cell(row, 4, '', fill=total_fill)
        cell(row, 5, teacher_total, font=total_font, fill=total_fill,
             align=center, num_fmt='0.0')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    # Grand total
    grand_total = float(TeacherSubject.objects.aggregate(
        s=models.Sum('hours_per_week')
    )['s'] or 0)
    gt_font = Font(bold=True, color='FFFFFF', size=11)
    gt_fill = PatternFill('solid', fgColor='2C5F8A')
    cell(row, 1, 'ЗАГАЛОМ', font=gt_font, fill=gt_fill, align=center)
    cell(row, 2, '', font=gt_font, fill=gt_fill)
    cell(row, 3, '', font=gt_font, fill=gt_fill)
    cell(row, 4, '', font=gt_font, fill=gt_fill)
    cell(row, 5, grand_total, font=gt_font, fill=gt_fill, align=center, num_fmt='0.0')
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    # ── Column widths ──
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="teacher_load.xlsx"'
    return resp


def export_class_load(request):
    """XLSX: all classes with subjects, teachers, groups and hours."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Навантаження по класах'

    # ── Styles ──
    hdr_font   = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill   = PatternFill('solid', fgColor='2C5F8A')
    total_font = Font(bold=True, size=10)
    total_fill = PatternFill('solid', fgColor='EAF4EA')
    center     = Alignment(horizontal='center', vertical='center')
    thin       = Side(style='thin', color='BFBFBF')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(row, col, value, font=None, fill=None, align=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:    c.font   = font
        if fill:    c.fill   = fill
        if align:   c.alignment = align
        if num_fmt: c.number_format = num_fmt
        c.border = border
        return c

    # ── Header ──
    headers = ['Клас', 'Предмет', 'Вчитель', 'Група', 'Год/тиж']
    for col, h in enumerate(headers, 1):
        cell(1, col, h, font=hdr_font, fill=hdr_fill, align=center)
    ws.row_dimensions[1].height = 20

    # ── Pre-compute merged total per class (grouped subjects counted once) ──
    # For each (class, subject) pair take the max hours across groups, then sum.
    from django.db.models import Max
    merged_by_class: dict = defaultdict(float)
    for row_agg in (TeacherSubject.objects
                    .values('school_class_id', 'subject_id')
                    .annotate(h=Max('hours_per_week'))):
        merged_by_class[row_agg['school_class_id']] += float(row_agg['h'])

    # ── Data ──
    assignments = (
        TeacherSubject.objects
        .select_related('teacher', 'subject', 'school_class')
        .order_by('school_class__grade', 'school_class__letter',
                  'subject__name', 'group', 'teacher__last_name')
    )

    merge2_fill = PatternFill('solid', fgColor='D5E8D4')  # slightly different green

    def write_class_subtotal(r, cls_name, cls_pk, raw_total):
        merged = merged_by_class.get(cls_pk, raw_total)
        # Row 1: full sum (groups counted separately)
        cell(r, 1, f'Разом (з групами): {cls_name}', font=total_font, fill=total_fill)
        for c in range(2, 5): cell(r, c, '', fill=total_fill)
        cell(r, 5, raw_total, font=total_font, fill=total_fill, align=center, num_fmt='0.0')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        # Row 2: merged sum (group duplicates removed)
        cell(r, 1, f'Разом (без груп): {cls_name}', font=total_font, fill=merge2_fill)
        for c in range(2, 5): cell(r, c, '', fill=merge2_fill)
        cell(r, 5, merged, font=total_font, fill=merge2_fill, align=center, num_fmt='0.0')
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        return r + 1

    row           = 2
    current_class = None
    current_cls_pk = None
    class_total   = 0

    for ts in assignments:
        cls_name = str(ts.school_class)

        if current_class is not None and cls_name != current_class:
            row = write_class_subtotal(row, current_class, current_cls_pk, class_total)
            class_total = 0

        current_class  = cls_name
        current_cls_pk = ts.school_class_id
        class_total   += ts.hours_per_week

        cell(row, 1, cls_name, align=center)
        cell(row, 2, ts.subject.name)
        cell(row, 3, str(ts.teacher))
        cell(row, 4, ts.group if ts.group else '', align=center)
        cell(row, 5, float(ts.hours_per_week), align=center, num_fmt='0.0')
        row += 1

    # Last class subtotal
    if current_class:
        row = write_class_subtotal(row, current_class, current_cls_pk, class_total)

    # Grand totals
    grand_total = float(TeacherSubject.objects.aggregate(
        s=models.Sum('hours_per_week')
    )['s'] or 0)
    grand_merged = sum(merged_by_class.values())
    gt_font = Font(bold=True, color='FFFFFF', size=11)
    gt_fill = PatternFill('solid', fgColor='2C5F8A')
    for label, value in [('ЗАГАЛОМ (з групами)', grand_total),
                          ('ЗАГАЛОМ (без груп)', grand_merged)]:
        cell(row, 1, label, font=gt_font, fill=gt_fill, align=center)
        for c in range(2, 5): cell(row, c, '', font=gt_font, fill=gt_fill)
        cell(row, 5, value, font=gt_font, fill=gt_fill, align=center, num_fmt='0.0')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    # ── Column widths ──
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = 'attachment; filename="class_load.xlsx"'
    return resp


# ─── Schedule XLSX export ─────────────────────────────────────────────────────

def export_schedule(request, pk):
    """XLSX: full schedule — one sheet per class, one sheet per teacher."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    schedule = get_object_or_404(Schedule, pk=pk)
    lessons = list(schedule.lessons
                   .select_related('school_class', 'subject', 'teacher', 'room')
                   .order_by('week', 'group'))
    classes = SchoolClass.objects.all().order_by('grade', 'letter')
    teacher_ids = {l.teacher_id for l in lessons}
    teachers = Teacher.objects.filter(pk__in=teacher_ids).order_by('last_name', 'first_name')
    D = schedule.days_per_week
    P = schedule.lessons_per_day
    days = DAYS_LABELS[:D]
    days_full = DAYS_FULL[:D]

    cls_grid = _build_display_grid(lessons, classes, D, P)

    # Teacher grid: {teacher_pk: {d: {p: cell_dict}}}
    def _teacher_grid(teacher):
        t_lessons = [l for l in lessons if l.teacher_id == teacher.pk]
        grid_t = {d: {p: [] for p in range(P)} for d in range(D)}
        for l in t_lessons:
            if l.day < D and l.period < P:
                grid_t[l.day][l.period].append(l)
        result = {}
        for d in range(D):
            result[d] = {}
            for p in range(P):
                cell_lessons = grid_t[d][p]
                if not cell_lessons:
                    result[d][p] = None
                    continue
                week_a = [l for l in cell_lessons if l.week == 0]
                week_b = [l for l in cell_lessons if l.week == 1]
                subj_a = {l.subject_id for l in week_a}
                subj_b = {l.subject_id for l in week_b}
                if (week_a and week_b
                        and subj_a == subj_b
                        and week_a[0].school_class_id == week_b[0].school_class_id):
                    result[d][p] = {'kind': 'regular', 'primary': week_a[0], 'extra': None}
                elif week_a and week_b:
                    result[d][p] = {'kind': 'alt', 'week_a': week_a[0], 'week_b': week_b[0]}
                else:
                    lesson = (week_a or week_b)[0]
                    result[d][p] = {'kind': 'regular', 'primary': lesson, 'extra': None}
        return result

    # Bell times
    bell_times = {}
    if schedule.bell_schedule_id:
        from .models import BellPeriod
        bell_times = {
            bp.number - 1: bp
            for bp in BellPeriod.objects.filter(bell_schedule_id=schedule.bell_schedule_id)
        }

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Colour palette ────────────────────────────────────────────────────────
    C_NAVY    = '1B3A6B'
    C_BLUE    = '2C5F8A'
    C_LT_BLUE = 'D6E4F0'
    C_WHITE   = 'FFFFFF'
    C_GRAY    = 'F0F0F0'
    C_ALT     = 'FFF9C4'   # light amber — alt lesson cell
    C_ALT_A   = '1B5E20'   # dark green text — week A label
    C_ALT_B   = 'BF360C'   # dark red-orange — week B label
    C_SUBJ    = '0D1B4B'   # dark navy — subject name
    C_INFO    = '444444'   # gray — teacher / room

    # ── Shared style helpers ──────────────────────────────────────────────────
    thin   = Side(style='thin',   color='BFBFBF')
    medium = Side(style='medium', color=C_BLUE)
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_hdr = Border(left=thin, right=thin, top=thin, bottom=Side(style='medium', color=C_NAVY))

    f_title   = Font(name='Calibri', bold=True, size=14, color=C_WHITE)
    f_day     = Font(name='Calibri', bold=True, size=11, color=C_WHITE)
    f_period  = Font(name='Calibri', bold=True, size=10, color=C_NAVY)
    f_subj    = InlineFont(rFont='Calibri', b=True,  sz=20, color=C_SUBJ)
    f_info    = InlineFont(rFont='Calibri', b=False, sz=18, color=C_INFO)
    f_alt_a   = InlineFont(rFont='Calibri', b=True,  sz=18, color=C_ALT_A)
    f_alt_b   = InlineFont(rFont='Calibri', b=True,  sz=18, color=C_ALT_B)
    f_dim     = InlineFont(rFont='Calibri', b=False, sz=16, color='999999')

    fill_navy  = PatternFill('solid', fgColor=C_NAVY)
    fill_blue  = PatternFill('solid', fgColor=C_BLUE)
    fill_lblue = PatternFill('solid', fgColor=C_LT_BLUE)
    fill_white = PatternFill('solid', fgColor=C_WHITE)
    fill_gray  = PatternFill('solid', fgColor=C_GRAY)

    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _hex_rgb(color_hex):
        """Parse hex color to (r, g, b) tuple. Returns None on failure."""
        try:
            h = (color_hex or '').lstrip('#')
            if len(h) != 6:
                return None
            return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
        except Exception:
            return None

    def _subject_fill(color_hex, alpha=0.22):
        """Blend subject hex color with white and return PatternFill."""
        rgb = _hex_rgb(color_hex)
        if not rgb:
            return fill_white
        r2 = round(rgb[0] * alpha + 255 * (1 - alpha))
        g2 = round(rgb[1] * alpha + 255 * (1 - alpha))
        b2 = round(rgb[2] * alpha + 255 * (1 - alpha))
        return PatternFill('solid', fgColor=f'{r2:02X}{g2:02X}{b2:02X}')

    def _two_subject_gradient(c1_hex, c2_hex, alpha=0.28):
        """Градієнт зверху-вниз: колір предмету гр.1 → колір предмету гр.2."""
        from openpyxl.styles.fills import GradientFill
        from openpyxl.styles.colors import Color

        def to_argb(h):
            rgb = _hex_rgb(h) or (180, 180, 180)
            r = round(rgb[0] * alpha + 255 * (1 - alpha))
            g = round(rgb[1] * alpha + 255 * (1 - alpha))
            b = round(rgb[2] * alpha + 255 * (1 - alpha))
            return f'FF{r:02X}{g:02X}{b:02X}'

        return GradientFill(
            type='linear', degree=90,
            stop=[Color(rgb=to_argb(c1_hex)), Color(rgb=to_argb(c2_hex))],
        )

    def _count_lines(val):
        """Count lines (explicit \\n + 1) in a cell value (str or CellRichText)."""
        if not val:
            return 1
        if isinstance(val, str):
            return val.count('\n') + 1
        n = 0
        for block in val:
            t = block.text if hasattr(block, 'text') else ''
            n += t.count('\n')
        return n + 1

    def _period_label(p):
        bp = bell_times.get(p)
        if bp:
            return f'{p + 1}\n{bp.start_time.strftime("%H:%M")}–{bp.end_time.strftime("%H:%M")}'
        return str(p + 1)

    def _lesson_rich(lesson):
        """Build CellRichText: bold subject (+ group if set), newline, teacher, newline, room."""
        subj = lesson.subject.short_name or lesson.subject.name
        group_prefix = f'Гр.{lesson.group}  ' if lesson.group else ''
        teacher_str = str(lesson.teacher)
        room_str = str(lesson.room) if lesson.room else ''
        f_grp = InlineFont(rFont='Calibri', b=True, sz=17, color='666666')
        parts = []
        if group_prefix:
            parts += [TextBlock(f_grp, group_prefix), TextBlock(f_subj, subj)]
        else:
            parts.append(TextBlock(f_subj, subj))
        parts.append(TextBlock(f_info, f'\n{teacher_str}'))
        if room_str:
            parts.append(TextBlock(f_dim, f'\n{room_str}'))
        return CellRichText(*parts)

    def _alt_rich(la, lb):
        """Build CellRichText for alt-week cell."""
        parts = []
        if la:
            sa = la.subject.short_name or la.subject.name
            ta = str(la.teacher)
            ra = str(la.room) if la.room else ''
            parts += [TextBlock(f_alt_a, 'А: '), TextBlock(f_subj, sa),
                      TextBlock(f_info, f'\n{ta}')]
            if ra:
                parts.append(TextBlock(f_dim, f'  {ra}'))
        if lb:
            sb = lb.subject.short_name or lb.subject.name
            tb = str(lb.teacher)
            rb = str(lb.room) if lb.room else ''
            prefix = '\n' if la else ''
            parts += [TextBlock(f_alt_b, f'{prefix}Б: '), TextBlock(f_subj, sb),
                      TextBlock(f_info, f'\n{tb}')]
            if rb:
                parts.append(TextBlock(f_dim, f'  {rb}'))
        return CellRichText(*parts) if parts else ''

    def _two_group_rich(primary, extra):
        """Two groups: як на вебі — однаковий предмет показуємо раз, різні — окремо з кольорами."""
        same_subj = primary.subject_id == extra.subject_id
        g1 = f'Гр.{primary.group}' if primary.group else 'Гр.1'
        g2 = f'Гр.{extra.group}' if extra.group else 'Гр.2'
        t1 = str(primary.teacher)
        r1 = str(primary.room) if primary.room else ''
        t2 = str(extra.teacher)
        r2 = str(extra.room) if extra.room else ''
        lf   = InlineFont(rFont='Calibri', b=False, sz=18, color=C_INFO)
        lf_g = InlineFont(rFont='Calibri', b=True,  sz=17, color='666666')

        if same_subj:
            # Предмет один — показуємо раз, потім рядок на кожну групу
            subj = primary.subject.short_name or primary.subject.name
            parts = [TextBlock(f_subj, subj)]
            parts += [TextBlock(lf_g, f'\n{g1}: '), TextBlock(lf, t1)]
            if r1:
                parts.append(TextBlock(f_dim, f'  {r1}'))
            parts += [TextBlock(lf_g, f'\n{g2}: '), TextBlock(lf, t2)]
            if r2:
                parts.append(TextBlock(f_dim, f'  {r2}'))
        else:
            # Різні предмети — кожна група своїм кольором (як на вебі)
            s1 = primary.subject.short_name or primary.subject.name
            s2 = extra.subject.short_name or extra.subject.name
            parts = [TextBlock(lf_g, f'{g1}  '), TextBlock(f_subj, s1),
                     TextBlock(lf, f'\n{t1}')]
            if r1:
                parts.append(TextBlock(f_dim, f'  {r1}'))
            parts += [TextBlock(lf_g, f'\n{g2}  '), TextBlock(f_subj, s2),
                      TextBlock(lf, f'\n{t2}')]
            if r2:
                parts.append(TextBlock(f_dim, f'  {r2}'))
        return CellRichText(*parts)

    def _write_sheet(ws, title, grid_data, col_label_fn):
        """
        Populate a sheet.
        col_label_fn(col_i) → label for each data column (0-based, after period col).
        grid_data: {d: {p: cell_dict}}
        """
        # Column widths
        ws.column_dimensions['A'].width = 14
        for i in range(D):
            ws.column_dimensions[get_column_letter(i + 2)].width = 36

        # Row 1 — title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=D + 1)
        c = ws.cell(row=1, column=1, value=title)
        c.font   = f_title
        c.fill   = fill_navy
        c.alignment = al_center
        ws.row_dimensions[1].height = 28

        # Row 2 — day headers
        hdr = ws.cell(row=2, column=1, value='Урок')
        hdr.font = f_day; hdr.fill = fill_blue; hdr.alignment = al_center; hdr.border = brd_hdr
        for i in range(D):
            c = ws.cell(row=2, column=i + 2, value=col_label_fn(i))
            c.font = f_day; c.fill = fill_blue; c.alignment = al_center; c.border = brd_hdr
        ws.row_dimensions[2].height = 20

        # Rows 3+ — periods
        LINE_H = 26  # points per logical line (font ~10pt + inter-line + cell padding)
        ROW_MIN = 68
        for p in range(P):
            row = p + 3

            # Period column
            pc = ws.cell(row=row, column=1, value=_period_label(p))
            pc.font = f_period; pc.fill = fill_lblue; pc.alignment = al_center; pc.border = brd

            max_lines = 1
            for d in range(D):
                col = d + 2
                cell_data = grid_data[d][p]
                c = ws.cell(row=row, column=col)
                c.border = brd
                c.alignment = al_left

                if cell_data is None:
                    c.fill = fill_gray
                elif cell_data['kind'] == 'regular':
                    primary = cell_data['primary']
                    extra   = cell_data.get('extra')
                    if extra and primary.subject_id != extra.subject_id:
                        c.fill = _two_subject_gradient(primary.subject.color, extra.subject.color)
                    else:
                        c.fill = _subject_fill(primary.subject.color)
                    c.value = (_two_group_rich(primary, extra)
                               if extra else _lesson_rich(primary))
                else:  # alt
                    la = cell_data.get('week_a')
                    lb = cell_data.get('week_b')
                    base_color = (la or lb).subject.color if (la or lb) else None
                    c.fill = _subject_fill(base_color, alpha=0.18)
                    c.value = _alt_rich(la, lb)

                max_lines = max(max_lines, _count_lines(c.value))

            ws.row_dimensions[row].height = max(max_lines * LINE_H, ROW_MIN)

    # ── Class sheets ──────────────────────────────────────────────────────────
    for sc in classes:
        if sc.pk not in cls_grid:
            continue
        ws = wb.create_sheet(title=str(sc))
        title = f'Розклад {sc}  ·  {schedule.name}'
        _write_sheet(
            ws, title,
            cls_grid[sc.pk],
            col_label_fn=lambda i: days_full[i],
        )
        ws.freeze_panes = 'B3'

    # ── Teacher sheets ────────────────────────────────────────────────────────
    for t in teachers:
        safe_name = str(t)[:31]  # sheet name ≤ 31 chars
        ws = wb.create_sheet(title=safe_name)
        tg = _teacher_grid(t)
        title = f'{t}  ·  {schedule.name}'
        _write_sheet(
            ws, title,
            tg,
            col_label_fn=lambda i: days_full[i],
        )
        ws.freeze_panes = 'B3'

    # ── Save & respond ────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_sched = schedule.name.replace(' ', '_')[:40]
    resp = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="schedule_{safe_sched}.xlsx"'
    return resp


# ─── Rooms XLSX export ────────────────────────────────────────────────────────

def export_rooms(request, pk):
    """XLSX: один аркуш на кожен кабінет з його розкладом уроків."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    schedule = get_object_or_404(Schedule, pk=pk)
    lessons = list(schedule.lessons
                   .select_related('school_class', 'subject', 'teacher', 'room')
                   .filter(room__isnull=False)
                   .order_by('week', 'group'))

    room_ids = {l.room_id for l in lessons}
    rooms = Room.objects.filter(pk__in=room_ids).prefetch_related('subjects').order_by('name')

    D = schedule.days_per_week
    P = schedule.lessons_per_day
    days_full = DAYS_FULL[:D]

    bell_times = {}
    if schedule.bell_schedule_id:
        from .models import BellPeriod
        bell_times = {
            bp.number - 1: bp
            for bp in BellPeriod.objects.filter(bell_schedule_id=schedule.bell_schedule_id)
        }

    # ── Стилі (ті самі, що в export_schedule) ────────────────────────────────
    C_NAVY  = '1B3A6B'
    C_BLUE  = '2C5F8A'
    C_LT_BLUE = 'D6E4F0'
    C_WHITE = 'FFFFFF'
    C_GRAY  = 'F0F0F0'
    C_SUBJ  = '0D1B4B'
    C_INFO  = '444444'

    thin    = Side(style='thin',   color='BFBFBF')
    brd     = Border(left=thin, right=thin, top=thin, bottom=thin)
    brd_hdr = Border(left=thin, right=thin, top=thin, bottom=Side(style='medium', color=C_NAVY))

    f_title  = Font(name='Calibri', bold=True, size=14, color=C_WHITE)
    f_day    = Font(name='Calibri', bold=True, size=11, color=C_WHITE)
    f_period = Font(name='Calibri', bold=True, size=10, color=C_NAVY)
    f_subj   = InlineFont(rFont='Calibri', b=True,  sz=20, color=C_SUBJ)
    f_info   = InlineFont(rFont='Calibri', b=False, sz=18, color=C_INFO)
    f_dim    = InlineFont(rFont='Calibri', b=False, sz=16, color='999999')
    f_alt_a  = InlineFont(rFont='Calibri', b=True,  sz=18, color='1B5E20')
    f_alt_b  = InlineFont(rFont='Calibri', b=True,  sz=18, color='BF360C')

    fill_navy  = PatternFill('solid', fgColor=C_NAVY)
    fill_blue  = PatternFill('solid', fgColor=C_BLUE)
    fill_lblue = PatternFill('solid', fgColor=C_LT_BLUE)
    fill_white = PatternFill('solid', fgColor=C_WHITE)
    fill_gray  = PatternFill('solid', fgColor=C_GRAY)

    al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    al_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def _hex_rgb(h):
        try:
            h = (h or '').lstrip('#')
            return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)) if len(h) == 6 else None
        except Exception:
            return None

    def _subject_fill(color_hex, alpha=0.22):
        rgb = _hex_rgb(color_hex)
        if not rgb:
            return fill_white
        r2 = round(rgb[0] * alpha + 255 * (1 - alpha))
        g2 = round(rgb[1] * alpha + 255 * (1 - alpha))
        b2 = round(rgb[2] * alpha + 255 * (1 - alpha))
        return PatternFill('solid', fgColor=f'{r2:02X}{g2:02X}{b2:02X}')

    def _count_lines(val):
        if not val:
            return 1
        if isinstance(val, str):
            return val.count('\n') + 1
        return sum(b.text.count('\n') if hasattr(b, 'text') else 0 for b in val) + 1

    def _period_label(p):
        bp = bell_times.get(p)
        if bp:
            return f'{p + 1}\n{bp.start_time.strftime("%H:%M")}–{bp.end_time.strftime("%H:%M")}'
        return str(p + 1)

    def _room_cell_rich(lessons_in_slot):
        """Rich text для слоту кабінету: предмет, клас, вчитель."""
        parts = []
        for i, l in enumerate(lessons_in_slot):
            subj = l.subject.short_name or l.subject.name
            cls_str = str(l.school_class)
            grp = f' гр.{l.group}' if l.group else ''
            teacher_str = str(l.teacher)
            prefix = '\n' if i else ''
            parts += [
                TextBlock(f_subj,  f'{prefix}{subj}'),
                TextBlock(f_info,  f'\n{cls_str}{grp}'),
                TextBlock(f_dim,   f'\n{teacher_str}'),
            ]
        return CellRichText(*parts) if parts else ''

    def _build_room_grid(room_pk):
        """Побудувати grid для одного кабінету.

        Сім'я з обома тижнями (0 і 1) → звичайний урок, без мітки А/Б.
        Сім'я тільки з тижнем 0 → alt-A; тільки з тижнем 1 → alt-B.
        """
        r_lessons = [l for l in lessons if l.room_id == room_pk]
        raw = {d: {p: [] for p in range(P)} for d in range(D)}
        for l in r_lessons:
            if l.day < D and l.period < P:
                raw[l.day][l.period].append(l)

        result = {}
        for d in range(D):
            result[d] = {}
            for p in range(P):
                slot = raw[d][p]
                if not slot:
                    result[d][p] = None
                    continue

                # Групуємо по сім'ї
                families: dict = {}
                for l in slot:
                    key = (l.teacher_id, l.school_class_id, l.subject_id, l.group)
                    families.setdefault(key, []).append(l)

                regular, alt_a, alt_b = [], [], []
                for fam in families.values():
                    weeks = {l.week for l in fam}
                    if 0 in weeks and 1 in weeks:
                        regular.append(fam[0])   # обидва тижні → звичайний, беремо один
                    elif 0 in weeks:
                        alt_a.append(next(l for l in fam if l.week == 0))
                    else:
                        alt_b.append(next(l for l in fam if l.week == 1))

                if alt_a or alt_b:
                    result[d][p] = {'kind': 'alt', 'regular': regular,
                                    'week_a': alt_a, 'week_b': alt_b}
                else:
                    result[d][p] = {'kind': 'regular', 'lessons': regular}
        return result

    def _write_room_sheet(ws, title, grid):
        ws.column_dimensions['A'].width = 14
        for i in range(D):
            ws.column_dimensions[get_column_letter(i + 2)].width = 36

        # Заголовок
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=D + 1)
        c = ws.cell(row=1, column=1, value=title)
        c.font = f_title; c.fill = fill_navy; c.alignment = al_center
        ws.row_dimensions[1].height = 28

        # Дні
        hdr = ws.cell(row=2, column=1, value='Урок')
        hdr.font = f_day; hdr.fill = fill_blue; hdr.alignment = al_center; hdr.border = brd_hdr
        for i in range(D):
            c = ws.cell(row=2, column=i + 2, value=days_full[i])
            c.font = f_day; c.fill = fill_blue; c.alignment = al_center; c.border = brd_hdr
        ws.row_dimensions[2].height = 20

        LINE_H = 26
        ROW_MIN = 68
        for p in range(P):
            row = p + 3
            pc = ws.cell(row=row, column=1, value=_period_label(p))
            pc.font = f_period; pc.fill = fill_lblue; pc.alignment = al_center; pc.border = brd

            max_lines = 1
            for d in range(D):
                col = d + 2
                cell_data = grid[d][p]
                c = ws.cell(row=row, column=col)
                c.border = brd; c.alignment = al_left

                if cell_data is None:
                    c.fill = fill_gray
                elif cell_data['kind'] == 'alt':
                    parts = []
                    all_lessons_in_cell = (cell_data['regular']
                                           + cell_data['week_a']
                                           + cell_data['week_b'])
                    base_color = all_lessons_in_cell[0].subject.color if all_lessons_in_cell else None
                    c.fill = _subject_fill(base_color, alpha=0.18)
                    # Звичайні (обидва тижні) — без мітки А/Б
                    for l in cell_data['regular']:
                        s = l.subject.short_name or l.subject.name
                        pfx = '\n' if parts else ''
                        parts += [TextBlock(f_subj, f'{pfx}{s}'),
                                  TextBlock(f_info, f'\n{l.school_class}'),
                                  TextBlock(f_dim,  f'\n{l.teacher}')]
                    # Alt-A
                    for l in cell_data['week_a']:
                        s = l.subject.short_name or l.subject.name
                        pfx = '\n' if parts else ''
                        parts += [TextBlock(f_alt_a, f'{pfx}А: '),
                                  TextBlock(f_subj,  s),
                                  TextBlock(f_info,  f'\n{l.school_class}'),
                                  TextBlock(f_dim,   f'\n{l.teacher}')]
                    # Alt-B
                    for l in cell_data['week_b']:
                        s = l.subject.short_name or l.subject.name
                        pfx = '\n' if parts else ''
                        parts += [TextBlock(f_alt_b, f'{pfx}Б: '),
                                  TextBlock(f_subj,  s),
                                  TextBlock(f_info,  f'\n{l.school_class}'),
                                  TextBlock(f_dim,   f'\n{l.teacher}')]
                    c.value = CellRichText(*parts) if parts else ''
                else:
                    slot_lessons = cell_data['lessons']
                    c.fill = _subject_fill(slot_lessons[0].subject.color)
                    c.value = _room_cell_rich(slot_lessons)

                max_lines = max(max_lines, _count_lines(c.value))
            ws.row_dimensions[row].height = max(max_lines * LINE_H, ROW_MIN)

    # ── Генерація аркушів ────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for room in rooms:
        grid = _build_room_grid(room.pk)
        safe_name = room.name[:31]
        ws = wb.create_sheet(title=safe_name)
        subj_str = ', '.join(s.name for s in room.subjects.all())
        title = f'Кабінет {room.name}' + (f' ({subj_str})' if subj_str else '') + f'  ·  {schedule.name}'
        _write_room_sheet(ws, title, grid)
        ws.freeze_panes = 'B3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_sched = schedule.name.replace(' ', '_')[:40]
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="rooms_{safe_sched}.xlsx"'
    return resp


def export_teacher_timetable(request, pk):
    """XLSX: зведена таблиця вчителів — рядки=вчителі, колонки=день×урок, в клітинці клас+кабінет."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    schedule = get_object_or_404(Schedule, pk=pk)
    D = schedule.days_per_week
    P = schedule.lessons_per_day

    lessons = list(schedule.lessons
                   .select_related('teacher', 'school_class', 'room')
                   .order_by('teacher__last_name', 'teacher__first_name', 'day', 'period', 'week'))

    # Збираємо вчителів що мають уроки
    teacher_ids_ordered = []
    seen = set()
    for l in lessons:
        if l.teacher_id not in seen:
            seen.add(l.teacher_id)
            teacher_ids_ordered.append(l.teacher_id)
    teachers_map = {t.pk: t for t in Teacher.objects.filter(pk__in=teacher_ids_ordered)}
    teachers = [teachers_map[tid] for tid in teacher_ids_ordered if tid in teachers_map]

    # grid[teacher_pk][d][p] → список уроків (тижень А і Б можуть бути обидва)
    grid = {t.pk: {d: {p: [] for p in range(P)} for d in range(D)} for t in teachers}
    for l in lessons:
        if l.teacher_id in grid and l.day < D and l.period < P:
            grid[l.teacher_id][l.day][l.period].append(l)

    def cell_text(lesson_list):
        """Повертає plain text: клас на першому рядку, 'каб. X' на другому."""
        if not lesson_list:
            return ''

        def fmt(l, prefix=''):
            lines = [f'{prefix}{l.school_class}']
            if l.room:
                lines.append(f'каб. {l.room}')
            return '\n'.join(lines)

        week_a = [l for l in lesson_list if l.week == 0]
        week_b = [l for l in lesson_list if l.week == 1]

        if week_a and week_b:
            if (week_a[0].school_class_id == week_b[0].school_class_id
                    and week_a[0].subject_id == week_b[0].subject_id):
                return fmt(week_a[0])
            return fmt(week_a[0], 'А: ') + '\n' + fmt(week_b[0], 'Б: ')

        return fmt((week_a or week_b)[0])

    # ── Стилі ──────────────────────────────────────────────────────────────────
    DAYS_UK = DAYS_FULL[:D]

    hdr_day_fill  = PatternFill('solid', fgColor='1F3864')
    hdr_per_fill  = PatternFill('solid', fgColor='2E5BA8')
    teacher_fill  = PatternFill('solid', fgColor='F2F2F2')
    border_thin   = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    align_center  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left    = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Розклад вчителів'

    # ── Рядок 1: назва розкладу ────────────────────────────────────────────────
    total_cols = 1 + D * P
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value=f'Розклад вчителів  ·  {schedule.name}')
    title_cell.font = Font(bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', fgColor='1F3864')
    title_cell.alignment = align_center

    # ── Рядок 2: дні (мердж по P колонок кожен) + заголовок «Вчитель» ─────────
    ws.cell(row=2, column=1, value='Вчитель').font = Font(bold=True, color='FFFFFF')
    ws.cell(row=2, column=1).fill = hdr_day_fill
    ws.cell(row=2, column=1).alignment = align_center
    ws.cell(row=2, column=1).border = border_thin

    for d in range(D):
        col_start = 2 + d * P
        col_end   = col_start + P - 1
        if P > 1:
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        c = ws.cell(row=2, column=col_start, value=DAYS_UK[d])
        c.font      = Font(bold=True, color='FFFFFF')
        c.fill      = hdr_day_fill
        c.alignment = align_center
        c.border    = border_thin

    # ── Рядок 3: номери уроків ─────────────────────────────────────────────────
    ws.cell(row=3, column=1, value='').border = border_thin
    ws.cell(row=3, column=1).fill = hdr_per_fill
    for d in range(D):
        for p in range(P):
            col = 2 + d * P + p
            c = ws.cell(row=3, column=col, value=p + 1)
            c.font      = Font(bold=True, color='FFFFFF')
            c.fill      = hdr_per_fill
            c.alignment = align_center
            c.border    = border_thin

    # ── Дані: рядок на вчителя ────────────────────────────────────────────────
    for row_idx, teacher in enumerate(teachers):
        row = 4 + row_idx
        # Чергуємо фон рядків
        row_fill = PatternFill('solid', fgColor='F2F2F2') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')

        c = ws.cell(row=row, column=1, value=str(teacher))
        c.font      = Font(bold=True)
        c.fill      = teacher_fill
        c.alignment = align_left
        c.border    = border_thin

        for d in range(D):
            for p in range(P):
                col  = 2 + d * P + p
                text = cell_text(grid[teacher.pk][d][p])
                c = ws.cell(row=row, column=col, value=text)
                c.fill      = row_fill
                c.alignment = align_center
                c.border    = border_thin

    # ── Ширини колонок ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 22
    for d in range(D):
        for p in range(P):
            col_letter = get_column_letter(2 + d * P + p)
            ws.column_dimensions[col_letter].width = 13

    # ── Висоти рядків ─────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16
    for row_idx in range(len(teachers)):
        ws.row_dimensions[4 + row_idx].height = 52

    ws.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = quote(f'Зведений розклад вчителів - {schedule.name}.xlsx')
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename=\"teacher_timetable.xlsx\"; filename*=UTF-8''{fname}"
    return resp


def export_class_timetable(request, pk):
    """XLSX: зведена таблиця класів — рядки=класи, колонки=день×урок, в клітинці предмет+вчитель."""
    import io
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    schedule = get_object_or_404(Schedule, pk=pk)
    D = schedule.days_per_week
    P = schedule.lessons_per_day

    lessons = list(schedule.lessons
                   .select_related('school_class', 'subject', 'teacher', 'room')
                   .order_by('school_class__grade', 'school_class__letter', 'day', 'period', 'week'))

    classes = SchoolClass.objects.order_by('grade', 'letter')

    # grid[class_pk][d][p] → список уроків
    grid = {sc.pk: {d: {p: [] for p in range(P)} for d in range(D)} for sc in classes}
    for l in lessons:
        if l.school_class_id in grid and l.day < D and l.period < P:
            grid[l.school_class_id][l.day][l.period].append(l)

    def cell_rich(lesson_list):
        if not lesson_list:
            return ''

        def fmt(l, prefix=''):
            subj = (l.subject.short_name or l.subject.name).upper()
            grp  = f' гр.{l.group}' if l.group else ''
            return f'{prefix}{subj}{grp}\n{l.teacher}'

        week_a = [l for l in lesson_list if l.week == 0]
        week_b = [l for l in lesson_list if l.week == 1]

        if week_a and week_b:
            if week_a[0].subject_id == week_b[0].subject_id and week_a[0].school_class_id == week_b[0].school_class_id:
                return fmt(week_a[0])
            return fmt(week_a[0], 'А: ') + '\n' + fmt(week_b[0], 'Б: ')

        seen = set()
        parts = []
        for l in lesson_list:
            key = (l.subject_id, l.teacher_id, l.week)
            if key not in seen:
                seen.add(key)
                parts.append(fmt(l))
        return '\n'.join(parts)

    # ── Стилі ─────────────────────────────────────────────────────────────────
    DAYS_UK = DAYS_FULL[:D]

    hdr_day_fill = PatternFill('solid', fgColor='1F3864')
    hdr_per_fill = PatternFill('solid', fgColor='2E5BA8')
    cls_fill     = PatternFill('solid', fgColor='F2F2F2')
    border_thin  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Розклад класів'

    total_cols = 1 + D * P

    # Рядок 1: заголовок
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value=f'Розклад класів  ·  {schedule.name}')
    c.font = Font(bold=True, size=13, color='FFFFFF')
    c.fill = hdr_day_fill
    c.alignment = align_center

    # Рядок 2: «Клас» + дні
    c = ws.cell(row=2, column=1, value='Клас')
    c.font = Font(bold=True, color='FFFFFF')
    c.fill = hdr_day_fill
    c.alignment = align_center
    c.border = border_thin

    for d in range(D):
        col_start = 2 + d * P
        col_end   = col_start + P - 1
        if P > 1:
            ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        c = ws.cell(row=2, column=col_start, value=DAYS_UK[d])
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = hdr_day_fill
        c.alignment = align_center
        c.border = border_thin

    # Рядок 3: номери уроків
    ws.cell(row=3, column=1).fill = hdr_per_fill
    ws.cell(row=3, column=1).border = border_thin
    for d in range(D):
        for p in range(P):
            col = 2 + d * P + p
            c = ws.cell(row=3, column=col, value=p + 1)
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = hdr_per_fill
            c.alignment = align_center
            c.border = border_thin

    # Дані
    for row_idx, sc in enumerate(classes):
        row = 4 + row_idx
        row_fill = PatternFill('solid', fgColor='F2F2F2') if row_idx % 2 == 0 else PatternFill('solid', fgColor='FFFFFF')

        c = ws.cell(row=row, column=1, value=str(sc))
        c.font = Font(bold=True)
        c.fill = cls_fill
        c.alignment = align_left
        c.border = border_thin

        for d in range(D):
            for p in range(P):
                col  = 2 + d * P + p
                text = cell_rich(grid[sc.pk][d][p])
                c = ws.cell(row=row, column=col, value=text)
                c.fill = row_fill
                c.alignment = align_center
                c.border = border_thin

    # Ширини
    ws.column_dimensions['A'].width = 10
    for d in range(D):
        for p in range(P):
            ws.column_dimensions[get_column_letter(2 + d * P + p)].width = 14

    # Висоти
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16
    for row_idx in range(len(list(classes))):
        ws.row_dimensions[4 + row_idx].height = 52

    ws.freeze_panes = 'B4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from urllib.parse import quote
    fname = quote(f'Зведений розклад класів - {schedule.name}.xlsx')
    resp = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename=\"class_timetable.xlsx\"; filename*=UTF-8''{fname}"
    return resp
